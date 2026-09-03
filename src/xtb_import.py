"""
xtb_import.py — XTB broker statement importer

Parses the "Cash Operations" sheet from XTB Excel exports and converts
rows into Negotium transaction format.

Comment patterns:
  Stock purchase: "OPEN BUY 4/4.138 @ 48.3060"  → 4 shares
                  "OPEN BUY 0.1367 @ 1462.60"   → 0.1367 shares
  Stock sell:     "CLOSE BUY 3.9657/14.7171 @ 123.3700" → 3.9657 shares
"""
from __future__ import annotations

import logging
import re
import warnings
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

import storage
import pandas as pd

import storage
import config as cfg_module
from ledger_core import get_all_transactions
from ticker_translate import translate_ticker

log = logging.getLogger(__name__)

SHARE_RE = re.compile(r"(?:OPEN|CLOSE)\s+BUY\s+([\d.]+)")
TRANSFER_RATE_RE = re.compile(r"Exchange rate:\s*([\d.]+)")
TRANSFER_CURRENCY_RE = re.compile(r"Currency conversion,\s*\w+\s+to\s+(\w+)")


def _parse_shares(comment: str | None) -> float | None:
    if not comment:
        return None
    m = SHARE_RE.search(comment)
    return float(m.group(1)) if m else None


def _parse_transfer_rate(comment: str | None) -> float | None:
    if not comment:
        return None
    m = TRANSFER_RATE_RE.search(comment)
    return float(m.group(1)) if m else None


def _parse_transfer_target(comment: str | None) -> str | None:
    if not comment:
        return None
    m = TRANSFER_CURRENCY_RE.search(comment)
    return m.group(1).upper() if m else None


# XTB statements occasionally ship with a minimal stylesheet that lacks the
# default ("Normal") cell style. openpyxl warns about it and falls back to its
# own defaults — harmless, so we filter just that message to keep the terminal
# (and the app log) clean. The warning is identical under all supported
# openpyxl versions; anchored to the exact wording openpyxl 3.1.x emits.
_OPENPYXL_NO_DEFAULT_STYLE = "Workbook contains no default style, apply openpyxl's default"


def _open_workbook(file_path: str | Path):
    """Try openpyxl first, fall back to pandas/calamine if the sheet is unreadable.

    The returned object (an openpyxl ``Workbook`` or a pandas ``ExcelFile``)
    is owned by the caller, which must always close it — see the ``finally``
    blocks in ``validate_xtb_file`` / ``parse_xtb_excel``.
    """
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message=_OPENPYXL_NO_DEFAULT_STYLE)
        try:
            return openpyxl.load_workbook(file_path, data_only=True), "openpyxl"
        except Exception:
            pass
        try:
            return pd.ExcelFile(file_path, engine="openpyxl"), "pandas"
        except Exception:
            pass
        return pd.ExcelFile(file_path, engine="calamine"), "calamine"


def validate_xtb_file(file_path: str | Path) -> tuple[bool, str]:
    try:
        wb, engine = _open_workbook(file_path)
    except Exception as e:
        return False, f"Cannot open file: {e}"

    try:
        if engine == "openpyxl":
            if "Cash Operations" not in wb.sheetnames:
                return False, "Missing 'Cash Operations' sheet."
            ws = wb["Cash Operations"]
            header_row = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if row[0] == "Type":
                    header_row = row
                    break
        else:
            if "Cash Operations" not in wb.sheet_names:
                return False, "Missing 'Cash Operations' sheet."
            df = wb.parse("Cash Operations", header=None, nrows=10)
            header_row = None
            for _, row in df.iterrows():
                if row.iloc[0] == "Type":
                    header_row = tuple(row)
                    break

        if not header_row:
            return False, "Cannot find column headers (Type, Ticker, ...) in Cash Operations."

        required = {"Type", "Ticker", "Amount", "Time"}
        actual = {str(c) for c in header_row if c and str(c) != "nan"}
        missing = required - actual
        if missing:
            return False, f"Missing columns: {', '.join(missing)}"

        return True, "Valid XTB statement."
    except Exception as e:
        return False, f"Error reading file: {e}"
    finally:
        try:
            wb.close()
        except Exception:
            pass


def parse_xtb_excel(file_path: str | Path, currency: str) -> list[dict]:
    currency = currency.upper()
    rules = cfg_module.load().get("ticker_rules", [])
    log.info("Parsing %s (currency=%s)", file_path, currency)

    wb, engine = _open_workbook(file_path)

    raw: list[dict] = []
    header: list[str] = []

    try:
        if engine == "openpyxl":
            ws = wb["Cash Operations"]
            rows_iter = ws.iter_rows(values_only=True)
            for i, row in enumerate(rows_iter):
                if not header and any(str(c).strip().lower() == "type" for c in row if c):
                    header = [str(c) if c else "" for c in row]
                    continue
                if header:
                    raw.append(row)
        else:
            df = wb.parse("Cash Operations", header=None)
            # Find header row to skip metadata rows
            header_idx = 0
            for idx, row in df.iterrows():
                if row.iloc[0] == "Type":
                    header_idx = idx + 1
                    header = [str(c) if c else "" for c in row]
                    break
            df.columns = df.iloc[header_idx - 1].values
            df = df.iloc[header_idx:].reset_index(drop=True)
            raw = [tuple(row) for _, row in df.iterrows()]
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # Build column index map from header (handles different XTB export layouts)
    col = {name.strip().lower(): i for i, name in enumerate(header) if name.strip()}
    log.info("Columns detected: %s", list(col.keys()))

    def _col(name: str) -> int | None:
        return col.get(name.lower())

    idx_type     = _col("type")
    idx_ticker   = _col("ticker")
    idx_time     = _col("time")
    idx_amount   = _col("amount")
    idx_comment  = _col("comment")

    skipped_no_type = 0
    skipped_no_amount = 0
    skipped_no_time = 0
    skipped_unknown_type = 0
    transactions: list[dict] = []
    seen_types: dict[str, int] = {}
    for row in raw:
        op_type     = row[idx_type]     if idx_type     is not None and idx_type     < len(row) else None
        ticker      = row[idx_ticker]   if idx_ticker   is not None and idx_ticker   < len(row) else None
        time_val    = row[idx_time]     if idx_time     is not None and idx_time     < len(row) else None
        amount      = row[idx_amount]   if idx_amount   is not None and idx_amount   < len(row) else None
        comment     = row[idx_comment]  if idx_comment  is not None and idx_comment  < len(row) else None

        if not op_type or op_type == "Total":
            skipped_no_type += 1
            continue
        if amount is None or (isinstance(amount, float) and pd.isna(amount)):
            skipped_no_amount += 1
            continue

        seen_types[op_type] = seen_types.get(op_type, 0) + 1

        if isinstance(time_val, str):
            try:
                time_val = datetime.fromisoformat(time_val)
            except ValueError:
                skipped_no_time += 1
                continue
        if not isinstance(time_val, datetime):
            skipped_no_time += 1
            continue

        date_str = time_val.strftime("%Y-%m-%d")
        entries: list[dict] = []

        if op_type == "Stock purchase":
            shares = _parse_shares(comment)
            if shares and shares > 0:
                entries.append({"ticker": translate_ticker(str(ticker), rules), "amount": round(shares, 8)})
                entries.append({"ticker": currency, "amount": round(float(amount), 8)})

        elif op_type == "Stock sell":
            shares = _parse_shares(comment)
            if shares and shares > 0:
                entries.append({"ticker": translate_ticker(str(ticker), rules), "amount": round(-shares, 8)})
                entries.append({"ticker": currency, "amount": round(float(amount), 8)})

        elif op_type in ("Deposit", "Withdrawal"):
            entries.append({"ticker": currency, "amount": round(float(amount), 8),
                            "account_operation": True})

        elif op_type == "Transfer":
            entries.append({"ticker": currency, "amount": round(float(amount), 8),
                            "account_operation": True})

        elif op_type == "Dividend":
            entries.append({"ticker": currency, "amount": round(float(amount), 8)})

        elif op_type in ("Free funds interest", "Free funds interest tax",
                          "Withholding tax"):
            entries.append({"ticker": currency, "amount": round(float(amount), 8)})

        if entries:
            transactions.append({"date": date_str, "entries": entries})

    log.info("Row types seen: %s", dict(seen_types))
    log.info("Skipped: %d no type/total, %d no amount, %d no valid time",
             skipped_no_type, skipped_no_amount, skipped_no_time)

    transactions.sort(key=lambda r: r["date"])

    merged: list[dict] = []
    for rec in transactions:
        if merged and merged[-1]["date"] == rec["date"]:
            merged[-1]["entries"].extend(rec["entries"])
        else:
            merged.append({"date": rec["date"], "entries": list(rec["entries"])})

    log.info("Parsed %d raw transactions, merged to %d daily records", len(transactions), len(merged))

    log.info("Final: %d records with %d total entries", len(merged), sum(len(r["entries"]) for r in merged))

    return merged



def _existing_entry_counts() -> dict[tuple[str, str, float], int]:
    from ledger_core import existing_entry_counts
    return existing_entry_counts()


def import_xtb(file_path: str | Path, currency: str) -> dict:
    log.info("=== XTB import: %s (currency=%s) ===", file_path, currency)
    valid, msg = validate_xtb_file(file_path)
    if not valid:
        log.error("Validation failed: %s", msg)
        return {"success": False, "error": msg}

    transactions = parse_xtb_excel(file_path, currency)

    # Include existing ledger balance so manually-added corporate-action
    # entries (split/spin-off) cover sells from the broker statement
    from ledger_core import get_all_transactions
    starting: dict[str, float] = {}
    for rec in get_all_transactions():
        for e in rec["entries"]:
            t = e["ticker"].upper()
            if t in storage.SUPPORTED_CURRENCIES:
                continue
            starting[t] = starting.get(t, 0.0) + float(e["amount"])

    # Auto-fix negative positions (e.g. corporate actions where the broker
    # statement doesn't record the acquisition of new shares).
    # Inserts a zero-cost buy on the date the position first went negative.
    from ledger_core import auto_fix_negative_positions
    fixed = auto_fix_negative_positions(transactions, starting_balance=starting)
    if fixed:
        log.info(
            "Auto-fixed %d negative position(s) from corporate action: %s",
            len(fixed),
            ", ".join(f"{t} ({n} shares)" for t, n, _ in fixed),
        )

    existing = _existing_entry_counts()

    imported = 0
    skipped = 0
    for rec in transactions:
        new_entries = []
        for e in rec["entries"]:
            key = (rec["date"], e["ticker"].upper(), round(float(e["amount"]), 8))
            if existing.get(key, 0) > 0:
                existing[key] -= 1
            else:
                new_entries.append(e)
        if new_entries:
            from ledger_core import add_transaction
            add_transaction(rec["date"], new_entries)
            imported += 1
        else:
            skipped += 1

    log.info("Result: %d imported, %d skipped (duplicates)", imported, skipped)
    return {"success": True, "imported": imported, "skipped": skipped}
