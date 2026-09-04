"""XTB importer — pytest suite (split from the original monolithic runner)."""

from __future__ import annotations

from pathlib import Path


def test_xtb_parse_shares(tmp: Path):
    """_parse_shares extracts share count from XTB comment patterns."""
    from xtb_import import _parse_shares

    assert _parse_shares("OPEN BUY 4/4.138 @ 48.3060") == 4.0
    assert _parse_shares("OPEN BUY 0.1367 @ 1462.60") == 0.1367
    assert _parse_shares("CLOSE BUY 3.9657/14.7171 @ 123.3700") == 3.9657
    assert _parse_shares("OPEN BUY 1 @ 107.00") == 1.0
    assert _parse_shares(None) is None
    assert _parse_shares("") is None
    assert _parse_shares("no match here") is None


def test_xtb_parse_transfer_rate(tmp: Path):
    """_parse_transfer_rate extracts exchange rate from transfer comment."""
    from xtb_import import _parse_transfer_rate

    comment = "Currency conversion, EUR to USD from TA: 52016471 to: 51963109, Exchange rate:1.159044"
    assert _parse_transfer_rate(comment) == 1.159044

    assert _parse_transfer_rate(None) is None
    assert _parse_transfer_rate("no rate here") is None


def test_xtb_parse_transfer_target(tmp: Path):
    """_parse_transfer_target extracts target currency from transfer comment."""
    from xtb_import import _parse_transfer_target

    comment = "Currency conversion, EUR to USD from TA: 52016471 to: 51963109, Exchange rate:1.159044"
    assert _parse_transfer_target(comment) == "USD"

    comment2 = "Currency conversion, PLN to EUR from TA: 53394664 to: 52016471, Exchange rate:0.23"
    assert _parse_transfer_target(comment2) == "EUR"

    assert _parse_transfer_target(None) is None
    assert _parse_transfer_target("no currency here") is None


def test_xtb_transfer_creates_source_entry(tmp: Path):
    """Transfer import creates the source currency entry (each file has its own side)."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    ws.append(["", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Transfer", "", "",
        "2026-06-01 10:47:30", -956, 1288183841,
        "Currency conversion, EUR to USD from TA: 52016471 to: 51963109, Exchange rate:1.159044",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_transfer.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    txns = parse_xtb_excel(str(xlsx_path), "EUR")
    assert len(txns) == 1

    entries = txns[0]["entries"]
    assert len(entries) == 1

    eur_entry = entries[0]
    assert eur_entry["ticker"] == "EUR"
    assert eur_entry["amount"] == -956
    assert eur_entry.get("account_operation") is True


def test_xtb_deposit_creates_account_operation(tmp: Path):
    """Deposit import creates entry with account_operation=True."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append(["Deposit", "", "", "2026-01-15 12:00:00", 5000, 111, "eWallet deposit"])

    del wb["Sheet"]
    xlsx_path = tmp / "test_deposit.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    txns = parse_xtb_excel(str(xlsx_path), "EUR")
    assert len(txns) == 1

    entries = txns[0]["entries"]
    assert len(entries) == 1
    assert entries[0]["ticker"] == "EUR"
    assert entries[0]["amount"] == 5000
    assert entries[0].get("account_operation") is True


def test_xtb_stock_purchase_creates_two_entries(tmp: Path):
    """Stock purchase creates share entry + currency outflow entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Stock purchase", "AAPL", "Apple",
        "2026-01-15 12:00:00", -1250.70, 222,
        "OPEN BUY 10/10 @ 125.07",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_buy.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    txns = parse_xtb_excel(str(xlsx_path), "EUR")
    assert len(txns) == 1

    entries = txns[0]["entries"]
    assert len(entries) == 2

    stock_entry = next(e for e in entries if e["ticker"] == "AAPL")
    cash_entry = next(e for e in entries if e["ticker"] == "EUR")

    assert stock_entry["amount"] == 10.0
    assert cash_entry["amount"] == -1250.70


def test_xtb_withholding_tax(tmp: Path):
    """Withholding tax creates a currency entry (no account_operation)."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Withholding tax", "FB2A.DE", "Meta",
        "2026-03-26 10:57:00", -0.76, 333,
        "FB2A.DE USD WHT 30%",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_wht.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    txns = parse_xtb_excel(str(xlsx_path), "EUR")
    assert len(txns) == 1

    entries = txns[0]["entries"]
    assert len(entries) == 1
    assert entries[0]["ticker"] == "EUR"
    assert entries[0]["amount"] == -0.76
    assert entries[0].get("account_operation") is None or entries[0].get("account_operation") is False


def test_xtb_dividend_import(tmp: Path):
    """Dividend rows from XTB statements are imported as cash credits."""
    from xtb_import import import_xtb
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Dividend", "AAPL", "AAPL",
        "2026-05-20 10:00:00", 12.50, 222,
        "Dividend payment",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_dividend.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    result = import_xtb(str(xlsx_path), "USD")
    assert result["success"] is True
    assert result["imported"] == 1

    from ledger_core import get_all_transactions
    records = get_all_transactions()
    assert len(records) == 1
    entries = records[0]["entries"]
    assert len(entries) == 1
    assert entries[0]["ticker"] == "USD"
    assert entries[0]["amount"] == 12.50
    assert entries[0].get("account_operation") is None


def test_xtb_negative_position_auto_fixed(tmp: Path):
    """Sell without prior buy is auto-fixed with a zero-cost acquisition.

    When a corporate action (split/spin-off) creates shares that the broker
    statement doesn't record, the import automatically inserts a zero-cost
    buy on the date the position first went negative. The user can later
    adjust the cost basis if needed.
    """
    from xtb_import import import_xtb
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Stock sell", "S2B.WA", "S2B",
        "2026-05-20 10:00:00", 193.50, 111,
        "CLOSE BUY 5.7315/5.7315 @ 33.76",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_negative.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    result = import_xtb(str(xlsx_path), "PLN")
    assert result["success"] is True
    # Should import the sell + auto-fix buy = 2 records
    assert result["imported"] == 2


def test_xtb_no_auto_fix_when_balance_covers(tmp: Path):
    """Auto-fix skipped when existing ledger balance already covers the sell.

    If the user has already manually added a corporate-action entry (buy at
    zero cost), the import should succeed without auto-fixing. This prevents
    double-counting when the ledger already has the shares.
    """
    from xtb_import import import_xtb
    from ledger_core import add_transaction
    import openpyxl

    # User manually adds the corporate-action acquisition first
    add_transaction("2026-05-19", [{"ticker": "S2B.WA", "amount": 5.7315}])

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Stock sell", "S2B.WA", "S2B",
        "2026-05-20 10:00:00", 193.50, 111,
        "CLOSE BUY 5.7315/5.7315 @ 33.76",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_ca_covered.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    result = import_xtb(str(xlsx_path), "PLN")
    assert result["success"] is True
    # Should import only the sell — no auto-fix needed
    assert result["imported"] == 1





def test_xtb_validate_missing_sheet(tmp: Path):
    """validate_xtb_file: file without 'Cash Operations' sheet fails."""
    from xtb_import import validate_xtb_file
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Positions"
    wb.active.append(["Type", "Ticker", "Amount"])
    p = tmp / "no_cash.xlsx"
    wb.save(p)
    wb.close()

    valid, msg = validate_xtb_file(p)
    assert valid is False
    assert "Cash Operations" in msg


def test_xtb_validate_missing_columns(tmp: Path):
    """validate_xtb_file: sheet missing required columns fails."""
    from xtb_import import validate_xtb_file
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker"])  # Missing Amount, Time
    p = tmp / "missing_cols.xlsx"
    wb.save(p)
    wb.close()

    valid, msg = validate_xtb_file(p)
    assert valid is False
    assert "Missing columns" in msg


def test_xtb_validate_valid(tmp: Path):
    """validate_xtb_file: valid XTB file passes."""
    from xtb_import import validate_xtb_file
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Deposit", "USD", 1000, "2023-01-03 10:00:00", ""])
    p = tmp / "valid.xlsx"
    wb.save(p)
    wb.close()

    valid, msg = validate_xtb_file(p)
    assert valid is True


def test_xtb_parse_dividend(tmp: Path):
    """parse_xtb_excel: dividend creates a currency entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Dividend", "AAPL.US", 15.50, "2023-06-15 10:00:00", "Dividend AAPL"])
    p = tmp / "div.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0]["amount"] == 15.50
    assert txns[0]["entries"][0]["ticker"] == "USD"


def test_xtb_parse_withdrawal(tmp: Path):
    """parse_xtb_excel: withdrawal creates account_operation entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Withdrawal", "USD", -500.0, "2023-07-01 10:00:00", ""])
    p = tmp / "withdraw.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0].get("account_operation") is True
    assert txns[0]["entries"][0]["amount"] == -500.0


def test_xtb_parse_interest(tmp: Path):
    """parse_xtb_excel: free funds interest creates currency entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Free funds interest", "USD", 2.35, "2023-06-30 10:00:00", ""])
    p = tmp / "interest.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0]["amount"] == 2.35
    assert txns[0]["entries"][0]["ticker"] == "USD"


def test_xtb_parse_withholding_tax(tmp: Path):
    """parse_xtb_excel: withholding tax creates currency entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Withholding tax", "USD", -3.10, "2023-06-15 10:00:00", ""])
    p = tmp / "wht.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0]["amount"] == -3.10


def test_xtb_parse_deposit(tmp: Path):
    """parse_xtb_excel: deposit creates account_operation entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Deposit", "USD", 5000.0, "2023-01-03 10:00:00", ""])
    p = tmp / "deposit.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0].get("account_operation") is True
    assert txns[0]["entries"][0]["amount"] == 5000.0


def test_xtb_parse_transfer(tmp: Path):
    """parse_xtb_excel: transfer creates account_operation entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Transfer", "USD", 2000.0, "2023-03-15 10:00:00", ""])
    p = tmp / "transfer.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0].get("account_operation") is True
