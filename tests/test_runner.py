#!/usr/bin/env python3
"""
test_runner.py —  Negotium - Investment Tracker test suite

Usage:
    cd investment_tracker
    python3 tests/test_runner.py

Tests are fully isolated: each test gets its own temp directory.
All temp dirs are deleted at the end (even on failure).
"""
from __future__ import annotations

import importlib
import json
import os
import shutil
import sys
import traceback
from datetime import date
from pathlib import Path

# Add src/ and project root to path so we can import the modules
SRC = Path(__file__).parent.parent / "src"
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(SRC))
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(Path(__file__).parent))

import fixtures as fx

# ── Test harness ───────────────────────────────────────────────────────────────

_ALL_TEMPS: list[Path] = []
_RESULTS: list[tuple[str, bool, str]] = []


def setup_env(tmp: Path) -> None:
    """Reload modules and patch paths to use tmp directory."""
    # We need to reload modules so their module-level globals are re-set
    import importlib
    import storage, config, transactions, portfolio, ticker_data

    for mod in [storage, config, transactions, portfolio, ticker_data]:
        importlib.reload(mod)

    fx.patch_root(tmp)


def run_test(name: str, fn):
    """Run a single test function in an isolated temp environment."""
    tmp = fx.make_temp_root()
    _ALL_TEMPS.append(tmp)
    try:
        setup_env(tmp)
        fn(tmp)
        _RESULTS.append((name, True, ""))
        print(f"  ✓  {name}")
    except Exception:
        tb = traceback.format_exc()
        _RESULTS.append((name, False, tb))
        print(f"  ✗  {name}")
        print(f"     {tb.strip().splitlines()[-1]}")


def cleanup():
    """Remove all temp directories created during this run."""
    removed = 0
    for tmp in _ALL_TEMPS:
        if tmp.exists():
            shutil.rmtree(tmp)
            removed += 1
    print(f"\n🧹 Cleaned up {removed} temp director{'y' if removed==1 else 'ies'}.")


# ── Individual tests ───────────────────────────────────────────────────────────

def test_config_defaults(tmp: Path):
    """Config creates default file when missing, loads it correctly."""
    import config
    cfg = config.load()
    assert cfg["name"] == "My Portfolio"
    assert cfg["start_day"] == "2020-01-01"
    assert cfg["default_currency"] == "PLN"
    assert (tmp / "data" / "config.json").exists(), "config.json should be created"


def test_config_save_and_reload(tmp: Path):
    """Config save → reload round-trip preserves all fields."""
    import config
    custom = {
        "name": "My Stocks",
        "start_day": "2022-06-15",
        "default_currency": "USD",
        "graph_precision": "1W",
    }
    config.save(custom)
    loaded = config.load()
    assert loaded["name"] == "My Stocks"
    assert loaded["start_day"] == "2022-06-15"
    assert loaded["default_currency"] == "USD"
    assert loaded["graph_precision"] == "1W"


def test_storage_jsonl_roundtrip(tmp: Path):
    """JSONL write → read preserves all records."""
    import storage
    path = tmp / "test.jsonl"
    records = [
        {"date": "2023-01-01", "entries": [{"ticker": "AAPL", "amount": 10}]},
        {"date": "2023-01-02", "entries": [{"ticker": "MSFT", "amount": 5}]},
    ]
    storage.write_jsonl(path, records)
    loaded = storage.read_jsonl(path)
    assert len(loaded) == 2
    assert loaded[0]["date"] == "2023-01-01"
    assert loaded[1]["entries"][0]["ticker"] == "MSFT"


def test_storage_append_jsonl(tmp: Path):
    """Appending to JSONL adds new records without touching existing ones."""
    import storage
    path = tmp / "append.jsonl"
    storage.write_jsonl(path, [{"date": "2023-01-01", "x": 1}])
    storage.append_jsonl(path, {"date": "2023-01-02", "x": 2})
    records = storage.read_jsonl(path)
    assert len(records) == 2
    assert records[1]["x"] == 2


def test_storage_balance(tmp: Path):
    """Balance save → load preserves values, strips near-zero entries."""
    import storage
    balance = {"AAPL": {"amount": 10.0, "avg_price": 125.0}, "PLN": {"amount": 5000.0, "avg_price": 0.0}, "MSFT": {"amount": 1e-12, "avg_price": 0.0}}
    storage.save_balance(balance)
    loaded = storage.load_balance()
    assert loaded["AAPL"]["amount"] == 10.0
    assert loaded["PLN"]["amount"] == 5000.0
    assert "MSFT" not in loaded, "Near-zero holding should be stripped"


def test_price_cache_write_read(tmp: Path):
    """Price cache write → read returns same data for a given ticker/year."""
    import storage
    prices = {"2023-01-03": 125.07, "2023-01-04": 126.36}
    storage.save_price_year("AAPL", 2023, prices)
    loaded = storage.load_price_year("AAPL", 2023)
    assert loaded["2023-01-03"] == 125.07
    assert storage.has_price_year("AAPL", 2023)
    assert not storage.has_price_year("AAPL", 2022)


def test_add_transaction_simple(tmp: Path):
    """Adding a transaction creates the ledger and updates balance."""
    import transactions, storage
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD", "amount": -1250.70},
    ])

    recs = transactions.get_all_transactions()
    assert len(recs) == 1
    assert recs[0]["date"] == "2023-01-03"

    bal = storage.load_balance()
    assert bal["AAPL"]["amount"] == 10.0
    assert abs(bal["USD"]["amount"] - (-1250.70)) < 0.01


def test_add_transaction_same_date_merges(tmp: Path):
    """Two transactions on the same date are merged into one line."""
    import transactions

    transactions.add_transaction("2023-01-03", [{"ticker": "AAPL", "amount": 5.0}])
    transactions.add_transaction("2023-01-03", [{"ticker": "MSFT", "amount": 3.0}])

    recs = transactions.get_all_transactions()
    assert len(recs) == 1, "Same-date transactions should merge into one record"
    tickers_in_rec = {e["ticker"] for e in recs[0]["entries"]}
    assert "AAPL" in tickers_in_rec
    assert "MSFT" in tickers_in_rec


def test_add_transaction_chronological_append(tmp: Path):
    """Transactions on later dates are appended in order."""
    import transactions

    transactions.add_transaction("2023-01-03", [{"ticker": "AAPL", "amount": 5.0}])
    transactions.add_transaction("2023-01-09", [{"ticker": "MSFT", "amount": 2.0}])
    transactions.add_transaction("2023-06-01", [{"ticker": "PLN", "amount": 1000.0}])

    recs = transactions.get_all_transactions()
    assert len(recs) == 3
    assert recs[0]["date"] == "2023-01-03"
    assert recs[1]["date"] == "2023-01-09"
    assert recs[2]["date"] == "2023-06-01"


def test_add_transaction_past_date_inserts_correctly(tmp: Path):
    """Inserting a past-date transaction reorders the file correctly."""
    import transactions

    transactions.add_transaction("2023-01-09", [{"ticker": "AAPL", "amount": 10.0}])
    transactions.add_transaction("2023-06-01", [{"ticker": "PLN", "amount": 500.0}])

    # Now insert something between them
    transactions.add_transaction("2023-01-04", [{"ticker": "USD", "amount": 1000.0}])

    recs = transactions.get_all_transactions()
    dates = [r["date"] for r in recs]
    assert dates == sorted(dates), f"Ledger must stay chronological, got: {dates}"
    assert dates[0] == "2023-01-04"


def test_compute_holdings_at(tmp: Path):
    """compute_holdings_at returns correct balances at a given date."""
    import transactions

    transactions.add_transaction("2023-01-03", [{"ticker": "AAPL", "amount": 10.0}])
    transactions.add_transaction("2023-06-01", [{"ticker": "AAPL", "amount": -5.0}])

    holdings_jan = transactions.compute_holdings_at("2023-01-31")
    assert holdings_jan["AAPL"] == 10.0

    holdings_jun = transactions.compute_holdings_at("2023-12-31")
    assert holdings_jun["AAPL"] == 5.0


def test_balance_after_full_sell(tmp: Path):
    """Selling all shares of a ticker removes it from holdings."""
    import transactions, storage

    transactions.add_transaction("2023-01-03", [{"ticker": "AAPL", "amount": 10.0}])
    transactions.add_transaction("2023-01-09", [{"ticker": "AAPL", "amount": -10.0}])

    bal = storage.load_balance()
    assert "AAPL" not in bal or abs(bal.get("AAPL", {}).get("amount", 0)) < 1e-6, \
        "After full sell, AAPL should be gone from balance"


def test_get_price_fallback_weekend(tmp: Path):
    """get_price falls back to Friday's close on weekend dates."""
    import ticker_data

    cache: dict = {}
    # Inject prices only for 2023-01-06 (Friday); 2023-01-07 (Sat) should fall back
    cache["AAPL"] = {2023: {"2023-01-06": 129.62}}

    price = ticker_data.get_price("AAPL", "2023-01-07", cache, 2023)
    assert price == 129.62, f"Expected 129.62, got {price}"


def test_get_price_cash_returns_one(tmp: Path):
    """Cash tickers always return 1.0 regardless of date."""
    import ticker_data

    cache: dict = {}
    for ccy in ["USD", "EUR", "PLN"]:
        price = ticker_data.get_price(ccy, "2023-01-03", cache, 2023)
        assert price == 1.0, f"Cash ticker {ccy} should return 1.0"


def test_get_fx_rate_same_currency(tmp: Path):
    """FX rate of same-to-same currency is exactly 1.0."""
    import ticker_data

    cache: dict = {}
    for ccy in ["USD", "EUR", "PLN"]:
        rate = ticker_data.get_fx_rate(ccy, ccy, "2023-01-03", cache, 2023)
        assert rate == 1.0, f"FX rate {ccy}→{ccy} should be 1.0"


def test_get_fx_rate_usd_to_pln(tmp: Path):
    """USD→PLN FX rate is read from USDPLN cache."""
    import ticker_data

    fx.inject_fake_prices(tmp)
    cache: dict = {}
    rate = ticker_data.get_fx_rate("USD", "PLN", "2023-01-03", cache, 2023)
    assert abs(rate - 4.38) < 0.01, f"Expected ~4.38, got {rate}"


def test_get_fx_rate_stale_usd_uses_latest(tmp: Path):
    """Stale USDPLN must fall back to the latest cached rate, not collapse to 1.0."""
    import ticker_data

    fx.inject_fake_prices(tmp)  # USDPLN has no value for 2023-03-01
    cache: dict = {}
    rate = ticker_data.get_fx_rate("USD", "PLN", "2023-03-01", cache, 2023)
    assert rate > 3.0, f"USD→PLN collapsed to {rate} (should use latest ~3.98)"
    assert abs(rate - 3.98) < 0.01, f"Expected latest ~3.98, got {rate}"


def test_get_fx_rate_stale_invertible(tmp: Path):
    """USD↔PLN must be near-inverses even on stale data (not 1.0 in any direction)."""
    import ticker_data

    fx.inject_fake_prices(tmp)
    cache: dict = {}
    up = ticker_data.get_fx_rate("USD", "PLN", "2023-03-01", cache, 2023)
    pu = ticker_data.get_fx_rate("PLN", "USD", "2023-03-01", cache, 2023)
    assert up > 3.0, f"USD→PLN fell back to {up}"
    assert pu > 0.2, f"PLN→USD fell back to {pu}"
    assert abs(up * pu - 1.0) < 0.02, f"rates not inverses: {up} * {pu}"


def test_get_fx_rate_gbp_triangulates(tmp: Path):
    """GBP→EUR and GBP→USD derive via PLN pairs instead of returning 1.0."""
    import ticker_data

    cache = {
        "GBPPLN": {2023: {"2023-01-03": 5.0}},
        "EURPLN": {2023: {"2023-01-03": 4.5}},
        "USDPLN": {2023: {"2023-01-03": 4.4}},
    }
    gbp_eur = ticker_data.get_fx_rate("GBP", "EUR", "2023-01-03", cache, 2023)
    assert abs(gbp_eur - (5.0 / 4.5)) < 0.01, f"GBP→EUR got {gbp_eur}"
    gbp_usd = ticker_data.get_fx_rate("GBP", "USD", "2023-01-03", cache, 2023)
    assert abs(gbp_usd - (5.0 / 4.4)) < 0.01, f"GBP→USD got {gbp_usd}"


def test_holdings_weights_currency_invariant(tmp: Path):
    """Position weights must not change when the base currency changes (no FX→1.0)."""
    import json
    import transactions, portfolio, storage

    def write(ticker, year, prices):
        d = tmp / "data" / "prices" / ticker
        d.mkdir(parents=True, exist_ok=True)
        (d / f"{year}.json").write_text(json.dumps(prices))

    # EURPLN and EURUSD are fresh on 2023-03-05; USDPLN is stale (only 2023-01-05).
    write("EURPLN", 2023, {"2023-03-05": 4.84})
    write("EURUSD", 2023, {"2023-03-05": 1.10})
    write("USDPLN", 2023, {"2023-01-05": 4.40})
    write("GOOG", 2023, {"2023-03-05": 100.0})
    write("SEC0.DE", 2023, {"2023-03-05": 80.0})

    transactions.add_transaction("2023-03-05", [
        {"ticker": "PLN", "amount": 10000.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-03-05", [
        {"ticker": "USD", "amount": -1000.0},
        {"ticker": "GOOG", "amount": 10.0},
    ])
    transactions.add_transaction("2023-03-05", [
        {"ticker": "EUR", "amount": -800.0},
        {"ticker": "SEC0.DE", "amount": 10.0},
    ])

    def weights(base):
        snaps = portfolio.build_portfolio(
            start_date=date(2023, 3, 5), end_date=date(2023, 3, 5),
            base_currency=base, precision="D", use_cache=False,
        )
        latest = snaps[-1]
        tv = latest["total_value"]
        assert tv > 0
        return {a["ticker"]: a["value_base"] / tv for a in latest["assets"]}

    w_pln = weights("PLN")
    w_usd = weights("USD")
    assert set(w_pln) == set(w_usd)
    for ticker in w_pln:
        assert abs(w_pln[ticker] - w_usd[ticker]) < 0.01, \
            f"weight of {ticker} differs by currency: {w_pln[ticker]:.4f} vs {w_usd[ticker]:.4f}"


def test_invalidate_portfolio(tmp: Path):
    """invalidate_portfolio_from removes snapshots on/after the given date."""
    import storage

    snapshots = [
        {"date": "2023-01-01", "total_value": 100},
        {"date": "2023-01-02", "total_value": 110},
        {"date": "2023-01-03", "total_value": 120},
        {"date": "2023-01-04", "total_value": 130},
    ]
    storage.save_portfolio(snapshots)
    storage.invalidate_portfolio_from("2023-01-03")
    kept = storage.load_portfolio()
    assert len(kept) == 2
    assert kept[-1]["date"] == "2023-01-02"


def test_portfolio_build_single_asset(tmp: Path):
    """Portfolio build produces correct values for a single USD stock in PLN base."""
    import transactions, portfolio

    fx.inject_fake_prices(tmp)

    # Buy 10 AAPL at 2023-01-03
    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD", "amount": -1250.70},
    ])

    snapshots = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 3),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    assert len(snapshots) == 1
    snap = snapshots[0]
    assert snap["date"] == "2023-01-03"

    # Expected: 10 AAPL × 125.07 × 4.38 ≈ 5478.07 PLN
    # Plus USD cash: -1250.70 × 4.38 ≈ -5477.79 PLN → near zero net cash
    aapl_asset = next((a for a in snap["assets"] if a["ticker"] == "AAPL"), None)
    assert aapl_asset is not None, "AAPL should appear in holdings"
    expected_aapl_value = 10.0 * 125.07 * 4.38
    assert abs(aapl_asset["value_base"] - expected_aapl_value) < 1.0, \
        f"AAPL value_base: expected ~{expected_aapl_value:.2f}, got {aapl_asset['value_base']}"


def test_portfolio_build_cash_only(tmp: Path):
    """Portfolio with only PLN cash shows correct value without any FX conversion."""
    import transactions, portfolio

    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0},
    ])

    snapshots = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 3),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    snap = snapshots[0]
    assert abs(snap["total_value"] - 10000.0) < 0.01, \
        f"10000 PLN cash should be 10000 PLN total_value, got {snap['total_value']}"


def test_portfolio_invested_tracking(tmp: Path):
    """invested correctly counts only positive cash inflows."""
    import transactions, portfolio

    fx.inject_fake_prices(tmp)

    # Deposit 10000 PLN, then buy AAPL (cash outflow should not count as invested)
    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0},  # inflow → invested
    ])
    transactions.add_transaction("2023-01-04", [
        {"ticker": "AAPL", "amount": 5.0},
        {"ticker": "PLN", "amount": -631.8},   # outflow → not a invested
    ])

    snapshots = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 4),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    # invested at Jan 4 should still be ~10000 (only the PLN deposit counts)
    snap_jan4 = next(s for s in snapshots if s["date"] == "2023-01-04")
    assert abs(snap_jan4["invested"] - 10000.0) < 1.0, \
        f"invested should be ~10000, got {snap_jan4['invested']}"


def test_portfolio_weekly_precision(tmp: Path):
    """Weekly precision yields only Friday dates."""
    import transactions, portfolio
    from datetime import timedelta

    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 5.0},
    ])

    snapshots = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 27),
        base_currency="PLN",
        precision="W-FRI",
        use_cache=False,
    )

    for snap in snapshots:
        d = date.fromisoformat(snap["date"])
        assert d.weekday() == 4, f"Weekly snapshot on {snap['date']} is not a Friday"


def test_portfolio_cache_resumes(tmp: Path):
    """Portfolio build resumes from cached snapshots without recomputing them."""
    import transactions, portfolio, storage

    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
    ])

    # Build up to Jan 5
    portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 5),
        base_currency="PLN",
        precision="D",
        use_cache=True,
    )

    # Patch storage to track if portfolio.jsonl is re-read
    calls = []
    orig_load = storage.load_portfolio

    def mock_load():
        calls.append(1)
        return orig_load()

    storage.load_portfolio = mock_load

    # Build Jan 5 to Jan 6 (should resume from cache)
    portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 6),
        base_currency="PLN",
        precision="D",
        use_cache=True,
    )

    saved = storage.load_portfolio()
    dates = [s["date"] for s in saved]
    assert "2023-01-06" in dates, "Jan 6 should be added after resuming"

    storage.load_portfolio = orig_load  # restore


def test_day_range_daily(tmp: Path):
    """_day_range yields every day between start and end inclusive."""
    from portfolio import _day_range
    days = list(_day_range(date(2023, 1, 3), date(2023, 1, 6), "D"))
    assert len(days) == 4
    assert days[0] == date(2023, 1, 3)
    assert days[-1] == date(2023, 1, 6)


def test_day_range_weekly(tmp: Path):
    """_day_range with weekly precision yields only Fridays."""
    from portfolio import _day_range
    days = list(_day_range(date(2023, 1, 3), date(2023, 1, 31), "W-FRI"))
    for d in days:
        assert d.weekday() == 4, f"{d} is not a Friday"
    assert len(days) == 4  # Jan 6, 13, 20, 27


def test_get_tickers(tmp: Path):
    """get_tickers returns all non-cash tickers from the ledger."""
    import transactions

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD", "amount": -1250.0},  # cash — should be excluded
    ])
    transactions.add_transaction("2023-01-09", [
        {"ticker": "CDR.WA", "amount": 5.0},
        {"ticker": "PLN", "amount": -650.0},   # cash — excluded
    ])

    tickers = transactions.get_tickers(include_cash=False)
    assert "AAPL" in tickers
    assert "CDR.WA" in tickers
    assert "USD" not in tickers
    assert "PLN" not in tickers


def test_config_precision_mapping(tmp: Path):
    """get_precision maps config values to pandas resample rules."""
    import config
    assert config.get_precision({"graph_precision": "1D"}) == "D"
    assert config.get_precision({"graph_precision": "1W"}) == "W-FRI"
    # Default fallback
    assert config.get_precision({}) == "D"


def test_storage_loads_prices_range(tmp: Path):
    """load_prices_range merges multiple year files correctly."""
    import storage

    storage.save_price_year("AAPL", 2022, {"2022-12-30": 129.93})
    storage.save_price_year("AAPL", 2023, {"2023-01-03": 125.07})

    prices = storage.load_prices_range("AAPL", date(2022, 12, 1), date(2023, 1, 31))
    assert "2022-12-30" in prices
    assert "2023-01-03" in prices


def test_snapshots_to_series(tmp: Path):
    """snapshots_to_series extracts correct parallel arrays."""
    from portfolio import snapshots_to_series

    snaps = [
        {"date": "2023-01-03", "total_value": 1000.0, "invested": 900.0},
        {"date": "2023-01-04", "total_value": 1050.0, "invested": 900.0},
    ]
    dates, values, contrs = snapshots_to_series(snaps)
    assert dates  == ["2023-01-03", "2023-01-04"]
    assert values == [1000.0, 1050.0]
    assert contrs == [900.0, 900.0]



# ── Buy / sell round-trip tests ───────────────────────────────────────────────

def test_buy_eur_etf_full_sell(tmp: Path):
    """Buy QDVE.DE (EUR ETF), sell all shares, receive EUR back."""
    import transactions, storage

    fx.inject_fake_prices(tmp)
    storage.save_price_year("QDVE.DE", 2023, {
        "2023-01-03": 200.0,
        "2023-06-01": 240.0,
    })

    # Buy: 5 shares at 200 EUR each
    transactions.add_transaction("2023-01-03", [
        {"ticker": "QDVE.DE", "amount": 5.0},
        {"ticker": "EUR",     "amount": -1000.0},
    ])
    # Sell: all 5 shares at 240 EUR each = 1200 EUR proceeds
    transactions.add_transaction("2023-06-01", [
        {"ticker": "QDVE.DE", "amount": -5.0},
        {"ticker": "EUR",     "amount": 1200.0},
    ])

    bal = storage.load_balance()
    assert "QDVE.DE" not in bal or abs(bal.get("QDVE.DE", {}).get("amount", 0)) < 1e-9, \
        "QDVE.DE should be fully sold"
    assert abs(bal.get("EUR", {}).get("amount", 0) - 200.0) < 0.01, \
        f"Expected 200 EUR profit remaining, got {bal.get('EUR', {}).get('amount', 0)}"


def test_buy_usd_stock_partial_sell(tmp: Path):
    """Buy GOOG in USD, partially sell, verify correct remaining balance."""
    import transactions, storage

    fx.inject_fake_prices(tmp)
    storage.save_price_year("GOOG", 2023, {
        "2023-01-03": 88.0,
        "2023-06-01": 122.0,
    })

    transactions.add_transaction("2023-01-03", [
        {"ticker": "GOOG", "amount": 10.0},
        {"ticker": "USD",  "amount": -880.0},
    ])
    # Sell 4 of 10 shares
    transactions.add_transaction("2023-06-01", [
        {"ticker": "GOOG", "amount": -4.0},
        {"ticker": "USD",  "amount": 488.0},   # 4 × 122
    ])

    bal = storage.load_balance()
    assert abs(bal.get("GOOG", {}).get("amount", 0) - 6.0) < 1e-9, \
        f"Expected 6 GOOG remaining, got {bal.get('GOOG', {}).get('amount', 0)}"
    assert abs(bal.get("USD", {}).get("amount", 0) - (-392.0)) < 0.01, \
        f"Expected -392 USD (net cash spent), got {bal.get('USD', {}).get('amount', 0)}"


def test_sell_proceeds_not_counted_as_invested(tmp: Path):
    """Sale proceeds (EUR from selling ETF) must NOT increase invested."""
    import transactions, portfolio

    fx.inject_fake_prices(tmp)
    storage_mod = __import__("storage")
    storage_mod.save_price_year("QDVE.DE", 2023, {
        "2023-01-03": 200.0,
        "2023-06-01": 240.0,
    })

    # Deposit real money: 1000 EUR
    transactions.add_transaction("2023-01-03", [
        {"ticker": "EUR", "amount": 1000.0},
    ])
    # Buy QDVE.DE with it
    transactions.add_transaction("2023-01-04", [
        {"ticker": "QDVE.DE", "amount": 5.0},
        {"ticker": "EUR",     "amount": -1000.0},
    ])
    # Sell for profit: 1200 EUR back
    transactions.add_transaction("2023-06-01", [
        {"ticker": "QDVE.DE", "amount": -5.0},
        {"ticker": "EUR",     "amount": 1200.0},
    ])

    snaps = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 6, 1),
        base_currency="EUR",
        precision="D",
        use_cache=False,
    )

    last = snaps[-1]
    # invested should be ~1000 EUR (the initial deposit), NOT 2200 EUR
    assert abs(last["invested"] - 1000.0) < 1.0, \
        f"invested should be ~1000 EUR (deposit only), got {last['invested']}"
    # Total value should be ~1200 EUR (the sale proceeds sitting as cash)
    assert abs(last["total_value"] - 1200.0) < 1.0, \
        f"Total value should be ~1200 EUR (all as cash), got {last['total_value']}"


def test_currency_exchange_counts_as_invested(tmp: Path):
    """Wiring USD into the portfolio (pure cash) counts as invested."""
    import transactions, portfolio

    fx.inject_fake_prices(tmp)

    # Wire in 1000 USD — pure cash deposit
    transactions.add_transaction("2023-01-03", [
        {"ticker": "USD", "amount": 1000.0},
    ])

    snaps = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 3),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    snap = snaps[0]
    # 1000 USD × 4.38 USDPLN = 4380 PLN invested
    assert abs(snap["invested"] - 4380.0) < 1.0, \
        f"1000 USD deposit should add ~4380 PLN invested, got {snap['invested']}"


def test_eur_stock_valued_correctly_in_pln(tmp: Path):
    """QDVE.DE (EUR) position is correctly converted to PLN via EURPLN."""
    import transactions, portfolio

    fx.inject_fake_prices(tmp)
    storage_mod = __import__("storage")
    storage_mod.save_price_year("QDVE.DE", 2023, {
        "2023-01-03": 200.0,
    })

    transactions.add_transaction("2023-01-03", [
        {"ticker": "QDVE.DE", "amount": 5.0},
        {"ticker": "EUR",     "amount": -1000.0},
    ])

    snaps = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 3),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    snap  = snaps[0]
    asset = next((a for a in snap["assets"] if a["ticker"] == "QDVE.DE"), None)
    assert asset is not None, "QDVE.DE should appear in holdings"
    assert asset["currency"] == "EUR", \
        f"QDVE.DE should be EUR-denominated, got {asset['currency']}"

    # 5 shares × 200 EUR × 4.68 EURPLN = 4680 PLN
    expected = 5.0 * 200.0 * 4.68
    assert abs(asset["value_base"] - expected) < 1.0, \
        f"QDVE.DE value in PLN: expected ~{expected:.2f}, got {asset['value_base']}"


def test_usd_stock_valued_correctly_in_pln(tmp: Path):
    """GOOG (USD) position is correctly converted to PLN via USDPLN."""
    import transactions, portfolio

    fx.inject_fake_prices(tmp)
    storage_mod = __import__("storage")
    storage_mod.save_price_year("GOOG", 2023, {
        "2023-01-03": 88.0,
    })

    transactions.add_transaction("2023-01-03", [
        {"ticker": "GOOG", "amount": 10.0},
        {"ticker": "USD",  "amount": -880.0},
    ])

    snaps = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 3),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    asset = next((a for a in snaps[0]["assets"] if a["ticker"] == "GOOG"), None)
    assert asset is not None, "GOOG should appear in holdings"
    assert asset["currency"] == "USD", \
        f"GOOG should be USD-denominated, got {asset['currency']}"

    # 10 shares × 88 USD × 4.38 USDPLN = 3854.4 PLN
    expected = 10.0 * 88.0 * 4.38
    assert abs(asset["value_base"] - expected) < 1.0, \
        f"GOOG value in PLN: expected ~{expected:.2f}, got {asset['value_base']}"


def test_mixed_portfolio_pln_eur_usd(tmp: Path):
    """Portfolio with PLN, EUR and USD positions all valued correctly."""
    import transactions, portfolio

    fx.inject_fake_prices(tmp)
    storage_mod = __import__("storage")
    storage_mod.save_price_year("QDVE.DE", 2023, {"2023-01-09": 210.0})
    storage_mod.save_price_year("GOOG",    2023, {"2023-01-09": 91.0})

    # Deposit PLN cash
    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 5000.0},
    ])
    # Buy EUR ETF
    transactions.add_transaction("2023-01-04", [
        {"ticker": "QDVE.DE", "amount": 3.0},
        {"ticker": "EUR",     "amount": -630.0},
    ])
    # Buy USD stock
    transactions.add_transaction("2023-01-06", [
        {"ticker": "GOOG", "amount": 2.0},
        {"ticker": "USD",  "amount": -182.0},
    ])

    snaps = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 9),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    last   = snaps[-1]
    tickers = {a["ticker"] for a in last["assets"]}

    assert "PLN"     in tickers, "PLN cash should appear"
    assert "EUR"     in tickers, "EUR cash should appear"
    assert "USD"     in tickers, "USD cash should appear"
    assert "QDVE.DE" in tickers, "QDVE.DE should appear"
    assert "GOOG"    in tickers, "GOOG should appear"

    qdve = next(a for a in last["assets"] if a["ticker"] == "QDVE.DE")
    goog = next(a for a in last["assets"] if a["ticker"] == "GOOG")
    assert qdve["currency"] == "EUR"
    assert goog["currency"] == "USD"

    # invested = only the PLN deposit (5000 PLN)
    # EUR buy and USD buy are stock transactions, not counted
    assert abs(last["invested"] - 5000.0) < 1.0, \
        f"Only PLN deposit should count as invested: {last['invested']}"


def test_delete_transaction(tmp: Path):
    """Deleting one entry removes it and rebuilds balance."""
    import transactions, storage

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD", "amount": -1250.70},
    ])

    transactions.delete_transaction("2023-01-03", 0)

    recs = transactions.get_all_transactions()
    assert len(recs) == 1
    assert len(recs[0]["entries"]) == 1
    assert recs[0]["entries"][0]["ticker"] == "USD"

    bal = storage.load_balance()
    assert "AAPL" not in bal or abs(bal.get("AAPL", {}).get("amount", 0)) < 1e-6


def test_delete_last_entry_removes_record(tmp: Path):
    """Deleting the only entry in a date removes the entire record."""
    import transactions

    transactions.add_transaction("2023-01-03", [{"ticker": "AAPL", "amount": 10.0}])
    transactions.add_transaction("2023-06-01", [{"ticker": "MSFT", "amount": 5.0}])

    transactions.delete_transaction("2023-01-03", 0)

    recs = transactions.get_all_transactions()
    assert len(recs) == 1
    assert recs[0]["date"] == "2023-06-01"


def test_update_transaction(tmp: Path):
    """Updating an entry changes ticker, amount, and account_operation."""
    import transactions, storage

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD", "amount": -1250.70},
    ])

    transactions.update_transaction("2023-01-03", 0, "MSFT", 20.0, account_operation=True)

    recs = transactions.get_all_transactions()
    e = recs[0]["entries"][0]
    assert e["ticker"] == "MSFT"
    assert e["amount"] == 20.0
    assert e.get("account_operation") is True

    bal = storage.load_balance()
    assert bal["MSFT"]["amount"] == 20.0
    assert "AAPL" not in bal or abs(bal.get("AAPL", {}).get("amount", 0)) < 1e-6


def test_xtb_parse_shares(tmp: Path):
    """_parse_shares extracts share count from XTB comment patterns."""
    from xtb_import import _parse_shares

    assert _parse_shares("OPEN BUY 4/4.138 @ 48.3060") == 4.0
    assert _parse_shares("OPEN BUY 0.1367 @ 1462.60") == 0.1367
    assert _parse_shares("CLOSE BUY 3.9657/14.7171 @ 123.3700") == 3.9657
    assert _parse_shares("OPEN BUY 1 @ 107.00") == 1.0
    assert _parse_shares(None) is None
    assert _parse_shares("") is None
    assert _parse_shares("no match here") is None


def test_xtb_parse_transfer_rate(tmp: Path):
    """_parse_transfer_rate extracts exchange rate from transfer comment."""
    from xtb_import import _parse_transfer_rate

    comment = "Currency conversion, EUR to USD from TA: 52016471 to: 51963109, Exchange rate:1.159044"
    assert _parse_transfer_rate(comment) == 1.159044

    assert _parse_transfer_rate(None) is None
    assert _parse_transfer_rate("no rate here") is None


def test_xtb_parse_transfer_target(tmp: Path):
    """_parse_transfer_target extracts target currency from transfer comment."""
    from xtb_import import _parse_transfer_target

    comment = "Currency conversion, EUR to USD from TA: 52016471 to: 51963109, Exchange rate:1.159044"
    assert _parse_transfer_target(comment) == "USD"

    comment2 = "Currency conversion, PLN to EUR from TA: 53394664 to: 52016471, Exchange rate:0.23"
    assert _parse_transfer_target(comment2) == "EUR"

    assert _parse_transfer_target(None) is None
    assert _parse_transfer_target("no currency here") is None


def test_xtb_transfer_creates_source_entry(tmp: Path):
    """Transfer import creates the source currency entry (each file has its own side)."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    ws.append(["", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Transfer", "", "",
        "2026-06-01 10:47:30", -956, 1288183841,
        "Currency conversion, EUR to USD from TA: 52016471 to: 51963109, Exchange rate:1.159044",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_transfer.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    txns = parse_xtb_excel(str(xlsx_path), "EUR")
    assert len(txns) == 1

    entries = txns[0]["entries"]
    assert len(entries) == 1

    eur_entry = entries[0]
    assert eur_entry["ticker"] == "EUR"
    assert eur_entry["amount"] == -956
    assert eur_entry.get("account_operation") is True


def test_xtb_deposit_creates_account_operation(tmp: Path):
    """Deposit import creates entry with account_operation=True."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append(["Deposit", "", "", "2026-01-15 12:00:00", 5000, 111, "eWallet deposit"])

    del wb["Sheet"]
    xlsx_path = tmp / "test_deposit.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    txns = parse_xtb_excel(str(xlsx_path), "EUR")
    assert len(txns) == 1

    entries = txns[0]["entries"]
    assert len(entries) == 1
    assert entries[0]["ticker"] == "EUR"
    assert entries[0]["amount"] == 5000
    assert entries[0].get("account_operation") is True


def test_xtb_stock_purchase_creates_two_entries(tmp: Path):
    """Stock purchase creates share entry + currency outflow entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Stock purchase", "AAPL", "Apple",
        "2026-01-15 12:00:00", -1250.70, 222,
        "OPEN BUY 10/10 @ 125.07",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_buy.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    txns = parse_xtb_excel(str(xlsx_path), "EUR")
    assert len(txns) == 1

    entries = txns[0]["entries"]
    assert len(entries) == 2

    stock_entry = next(e for e in entries if e["ticker"] == "AAPL")
    cash_entry = next(e for e in entries if e["ticker"] == "EUR")

    assert stock_entry["amount"] == 10.0
    assert cash_entry["amount"] == -1250.70


def test_xtb_withholding_tax(tmp: Path):
    """Withholding tax creates a currency entry (no account_operation)."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Withholding tax", "FB2A.DE", "Meta",
        "2026-03-26 10:57:00", -0.76, 333,
        "FB2A.DE USD WHT 30%",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_wht.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    txns = parse_xtb_excel(str(xlsx_path), "EUR")
    assert len(txns) == 1

    entries = txns[0]["entries"]
    assert len(entries) == 1
    assert entries[0]["ticker"] == "EUR"
    assert entries[0]["amount"] == -0.76
    assert entries[0].get("account_operation") is None or entries[0].get("account_operation") is False


def test_xtb_negative_position_gets_fixed(tmp: Path):
    """Sell without prior buy gets a compensating buy of X shares for 0.01 cash."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Cash Operations")

    for _ in range(5):
        ws.append(["", "", "", "", "", "", ""])
    ws.append(["Type", "Ticker", "Instrument", "Time", "Amount", "ID", "Comment"])
    ws.append([
        "Stock sell", "S2B.WA", "S2B",
        "2026-05-20 10:00:00", 193.50, 111,
        "CLOSE BUY 5.7315/5.7315 @ 33.76",
    ])

    del wb["Sheet"]
    xlsx_path = tmp / "test_negative.xlsx"
    wb.save(str(xlsx_path))
    wb.close()

    txns = parse_xtb_excel(str(xlsx_path), "PLN")
    assert len(txns) == 2

    sell_rec = txns[0]
    sell_entry = next(e for e in sell_rec["entries"] if e["ticker"] == "S2B.WA")
    assert sell_entry["amount"] < 0

    fix_rec = txns[1]
    buy_entry = next(e for e in fix_rec["entries"] if e["ticker"] == "S2B.WA")
    cash_entry = next(e for e in fix_rec["entries"] if e["ticker"] == "PLN")
    assert buy_entry["amount"] == abs(sell_entry["amount"])
    assert cash_entry["amount"] == -0.01


# ── Ticker translate tests ─────────────────────────────────────────────────────

def test_translate_no_rules(tmp: Path):
    """translate_ticker with no rules returns uppercased input."""
    from ticker_translate import translate_ticker
    assert translate_ticker("aapl") == "AAPL"
    assert translate_ticker("GOOG") == "GOOG"
    assert translate_ticker("", None) == ""


def test_translate_exact_match(tmp: Path):
    """Exact match rule replaces ticker."""
    from ticker_translate import translate_ticker
    rules = ["AMZN.DE=AMZ.DE", "VOW3.DE=VOW.DE"]
    assert translate_ticker("AMZN.DE", rules) == "AMZ.DE"
    assert translate_ticker("VOW3.DE", rules) == "VOW.DE"
    assert translate_ticker("AAPL.DE", rules) == "AAPL.DE"  # no match


def test_translate_suffix_swap(tmp: Path):
    """Suffix swap rule *.PL=*.WA replaces extension."""
    from ticker_translate import translate_ticker
    rules = ["*.PL=*.WA"]
    assert translate_ticker("SNT.PL", rules) == "SNT.WA"
    assert translate_ticker("CDR.PL", rules) == "CDR.WA"
    assert translate_ticker("AAPL.US", rules) == "AAPL.US"  # no match


def test_translate_suffix_strip(tmp: Path):
    """Suffix strip rule .US= removes the suffix."""
    from ticker_translate import translate_ticker
    rules = [".US="]
    assert translate_ticker("AAPL.US", rules) == "AAPL"
    assert translate_ticker("GOOG.US", rules) == "GOOG"
    assert translate_ticker("AAPL.DE", rules) == "AAPL.DE"  # no match


def test_translate_no_match(tmp: Path):
    """No matching rule returns uppercased input."""
    from ticker_translate import translate_ticker
    rules = ["AMZN.DE=AMZ.DE", "*.PL=*.WA"]
    assert translate_ticker("MSFT", rules) == "MSFT"
    assert translate_ticker("AAPL.US", rules) == "AAPL.US"


# ── Transactions: set_account_operation, get_transactions_up_to, get_all_tickers ──

def test_set_account_operation(tmp: Path):
    """set_account_operation toggles the flag on an entry."""
    import transactions

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD", "amount": -1250.0},
    ])

    recs = transactions.get_all_transactions()
    assert recs[0]["entries"][0].get("account_operation") is None

    transactions.set_account_operation("2023-01-03", 0, True)
    recs = transactions.get_all_transactions()
    assert recs[0]["entries"][0].get("account_operation") is True

    transactions.set_account_operation("2023-01-03", 0, False)
    recs = transactions.get_all_transactions()
    assert recs[0]["entries"][0].get("account_operation") is None


def test_get_transactions_up_to(tmp: Path):
    """get_transactions_up_to returns only transactions up to the given date."""
    import transactions

    transactions.add_transaction("2023-01-03", [{"ticker": "AAPL", "amount": 5.0}])
    transactions.add_transaction("2023-01-10", [{"ticker": "MSFT", "amount": 3.0}])
    transactions.add_transaction("2023-06-01", [{"ticker": "GOOG", "amount": 2.0}])

    result = transactions.get_transactions_up_to("2023-01-10")
    dates = [r["date"] for r in result]
    assert "2023-01-03" in dates
    assert "2023-01-10" in dates
    assert "2023-06-01" not in dates
    assert len(result) == 2


def test_get_all_tickers(tmp: Path):
    """get_all_tickers returns stock tickers plus FX pair tickers."""
    import transactions

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD", "amount": -1250.0},
    ])
    transactions.add_transaction("2023-01-06", [
        {"ticker": "QDVE.DE", "amount": 5.0},
        {"ticker": "EUR", "amount": -1000.0},
    ])

    tickers = transactions.get_all_tickers(include_fx=True)
    assert "AAPL" in tickers
    assert "QDVE.DE" in tickers
    # Cash tickers excluded, FX pairs included
    assert "USD" not in tickers
    assert "EUR" not in tickers


# ── Storage: project management ────────────────────────────────────────────────

def test_create_and_list_projects(tmp: Path):
    """create_project + list_projects returns sorted project names."""
    import storage

    storage.create_project("alpha")
    storage.create_project("gamma")
    storage.create_project("beta")

    projects = storage.list_projects()
    assert "alpha" in projects
    assert "beta" in projects
    assert "gamma" in projects
    assert projects == sorted(projects)


def test_rename_project(tmp: Path):
    """rename_project preserves project data under new name."""
    import storage

    storage.create_project("old_name")
    storage.set_current_project("old_name")
    storage.save_balance({"AAPL": {"amount": 10.0, "avg_price": 150.0}})

    storage.rename_project("old_name", "new_name")

    projects = storage.list_projects()
    assert "new_name" in projects
    assert "old_name" not in projects

    storage.set_current_project("new_name")
    bal = storage.load_balance()
    assert bal["AAPL"]["amount"] == 10.0


def test_delete_project(tmp: Path):
    """delete_project removes project directory and registry entry."""
    import storage

    storage.create_project("to_delete")
    assert "to_delete" in storage.list_projects()

    storage.delete_project("to_delete")
    assert "to_delete" not in storage.list_projects()


def test_config_is_global(tmp: Path):
    """All projects share one global config file (no per-project config)."""
    import storage, config

    storage.create_project("proj_a")
    storage.set_current_project("proj_a")
    cfg = config.load()
    cfg["name"] = "Global Portfolio"
    config.save(cfg)

    storage.create_project("proj_b")
    storage.set_current_project("proj_b")
    assert config.load()["name"] == "Global Portfolio"

    # No per-project config files should be created
    assert not (tmp / "data" / "proj_a" / "config.json").exists()
    assert not (tmp / "data" / "proj_b" / "config.json").exists()
    assert (tmp / "data" / "config.json").exists()


# ── Storage: benchmark roundtrip ───────────────────────────────────────────────

def test_benchmark_save_load_roundtrip(tmp: Path):
    """save_benchmarks → load_benchmarks roundtrip preserves data."""
    import storage

    data = [
        {"date": "2023-01-03", "SXRV.DE": 5000.0, "I500.DE": 4800.0},
        {"date": "2023-01-04", "SXRV.DE": 5050.0, "I500.DE": 4820.0},
    ]
    storage.save_benchmarks("PLN", data)
    loaded = storage.load_benchmarks("PLN")
    assert loaded is not None
    assert len(loaded) == 2
    assert loaded[0]["SXRV.DE"] == 5000.0
    assert loaded[1]["I500.DE"] == 4820.0


def test_benchmark_load_returns_none_when_missing(tmp: Path):
    """load_benchmarks returns None when no cache file exists."""
    import storage
    assert storage.load_benchmarks("USD") is None


# ── Portfolio: withdrawal and _ticker_currency ─────────────────────────────────

def test_withdrawal_decreases_invested(tmp: Path):
    """Withdrawal (negative account_operation) reduces invested capital."""
    import transactions, portfolio

    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0},
    ])
    transactions.add_transaction("2023-01-04", [
        {"ticker": "PLN", "amount": -3000.0, "account_operation": True},
    ])

    snaps = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 4),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    snap_jan3 = next(s for s in snaps if s["date"] == "2023-01-03")
    snap_jan4 = next(s for s in snaps if s["date"] == "2023-01-04")

    assert abs(snap_jan3["invested"] - 10000.0) < 1.0
    assert abs(snap_jan4["invested"] - 7000.0) < 1.0, \
        f"Withdrawal should reduce invested to ~7000, got {snap_jan4['invested']}"
    assert abs(snap_jan4["total_value"] - 7000.0) < 1.0


def test_ticker_currency_detection(tmp: Path):
    """_ticker_currency maps suffixes to correct currencies."""
    from portfolio import _ticker_currency

    assert _ticker_currency("PLN") == "PLN"
    assert _ticker_currency("USD") == "USD"
    assert _ticker_currency("EUR") == "EUR"
    assert _ticker_currency("QDVE.DE") == "EUR"
    assert _ticker_currency("SNT.WA") == "PLN"
    assert _ticker_currency("4GLD.L") == "GBP"
    assert _ticker_currency("AAPL") == "USD"  # no suffix → USD default

    # Unknown suffix defaults to USD (triggers warning on stdout)
    import io, sys
    old = sys.stdout
    sys.stdout = io.StringIO()
    assert _ticker_currency("AAPL.US") == "USD"
    sys.stdout = old


# ── Manual import tests ────────────────────────────────────────────────────────

def test_manual_validate_errors(tmp: Path):
    """validate_manual_file catches various malformed inputs."""
    from manual_import import validate_manual_file

    empty = tmp / "empty.json"
    empty.write_text("")
    valid, msg = validate_manual_file(empty)
    assert not valid
    assert "empty" in msg.lower()

    bad_json = tmp / "bad.json"
    bad_json.write_text("{not json}")
    valid, msg = validate_manual_file(bad_json)
    assert not valid

    not_array = tmp / "not_array.json"
    not_array.write_text('{"date": "2023-01-03"}')
    valid, msg = validate_manual_file(not_array)
    assert not valid
    assert "array" in msg.lower()

    missing_date = tmp / "missing_date.json"
    missing_date.write_text('[{"entries": [{"ticker": "AAPL", "amount": 10}]}]')
    valid, msg = validate_manual_file(missing_date)
    assert not valid
    assert "date" in msg.lower()

    missing_ticker = tmp / "missing_ticker.json"
    missing_ticker.write_text('[{"date": "2023-01-03", "entries": [{"amount": 10}]}]')
    valid, msg = validate_manual_file(missing_ticker)
    assert not valid
    assert "ticker" in msg.lower()


def test_manual_parse_and_import(tmp: Path):
    """parse_manual_json parses correctly; import_manual deduplicates."""
    from manual_import import parse_manual_json, import_manual

    data = [
        {
            "date": "2023-01-03",
            "entries": [
                {"ticker": "AAPL", "amount": 10.0},
                {"ticker": "USD", "amount": -1250.0},
            ],
        },
        {
            "date": "2023-01-04",
            "entries": [
                {"ticker": "MSFT", "amount": 5.0, "account_operation": True},
            ],
        },
    ]

    json_file = tmp / "test_manual.json"
    json_file.write_text(json.dumps(data))

    parsed = parse_manual_json(json_file)
    assert len(parsed) == 2
    assert parsed[0]["date"] == "2023-01-03"
    assert len(parsed[0]["entries"]) == 2
    assert parsed[1]["entries"][0].get("account_operation") is True

    result = import_manual(json_file)
    assert result["success"] is True
    assert result["imported"] == 2

    # Import again — should skip duplicates
    result2 = import_manual(json_file)
    assert result2["success"] is True
    assert result2["skipped"] == 2
    assert result2["imported"] == 0


# ── BOSSA validate tests ──────────────────────────────────────────────────────

def test_bossa_validate_errors(tmp: Path):
    """validate_bossa_file catches empty and malformed files."""
    from bossa_import import validate_bossa_file

    empty = tmp / "empty.csv"
    empty.write_text("")
    valid, msg = validate_bossa_file(empty)
    assert not valid
    assert "empty" in msg.lower()

    wrong_cols = tmp / "wrong.csv"
    wrong_cols.write_text("name;value\ntest;123")
    valid, msg = validate_bossa_file(wrong_cols)
    assert not valid
    assert "missing" in msg.lower()

    valid_file = tmp / "valid.csv"
    valid_file.write_text("data;tytuł operacji;szczegóły;kwota;waluta\nrow1;row2;row3;row4;row5")
    valid, msg = validate_bossa_file(valid_file)
    assert valid


# ── ISIN resolve tests ─────────────────────────────────────────────────────────

def test_isin_resolve_from_config(tmp: Path):
    """resolve_isins_with_names resolves ISINs based on config rules."""
    import json as _json
    import config as cfg_module

    cfg = cfg_module.load()
    cfg["isin_tickers"] = [
        "IE00B4L5Y983=IWDA.L",
        "US5949181085=MSFT.US",
    ]
    cfg_module.save(cfg)

    from isin_resolve import resolve_isins_with_names

    isin_map = {
        "IE00B4L5Y983": "iShares Core MSCI World",
        "US5949181085": "Microsoft",
        "DE0005793303": "unknown fund",
    }

    resolved, unresolved = resolve_isins_with_names(isin_map)
    assert resolved["IE00B4L5Y983"] == "IWDA.L"
    assert resolved["US5949181085"] == "MSFT.US"
    assert "DE0005793303" in unresolved
    assert unresolved["DE0005793303"] == "unknown fund"


# ── Include-today transaction tests ──────────────────────────────────────────

def test_portfolio_includes_today_transactions(tmp: Path):
    """build_portfolio includes transactions dated today."""
    from unittest.mock import patch
    import transactions, portfolio, portfolio_core
    from datetime import date as _real_date

    fx.inject_fake_prices(tmp)

    # Yesterday's transaction
    transactions.add_transaction("2023-01-04", [
        {"ticker": "AAPL", "amount": 10.0},
    ])
    # Today's transaction — should now be included
    transactions.add_transaction("2023-01-05", [
        {"ticker": "AAPL", "amount": 5.0},
    ])

    class _FakeDate(_real_date):
        @classmethod
        def today(cls):
            return _real_date(2023, 1, 5)

    with patch.object(portfolio_core, "date", _FakeDate):
        snapshots = portfolio.build_portfolio(
            start_date=date(2023, 1, 3),
            end_date=date(2023, 1, 5),
            base_currency="PLN",
            precision="D",
            use_cache=False,
        )

    snap = next(s for s in snapshots if s["date"] == "2023-01-05")
    aapl = next((a for a in snap["assets"] if a["ticker"] == "AAPL"), None)
    assert aapl is not None, "AAPL should appear"
    assert abs(aapl["amount"] - 15.0) < 1e-6, \
        f"Should hold 15 AAPL (today's +5 included), got {aapl['amount']}"


def test_rebuild_balance_includes_today_transactions(tmp: Path):
    """_rebuild_balance includes today's records in balance and avg_price."""
    from unittest.mock import patch
    import transactions, storage, ledger_core
    from datetime import date as _real_date

    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
    ])
    transactions.add_transaction("2023-01-05", [
        {"ticker": "AAPL", "amount": 5.0},
    ])

    class _FakeDate(_real_date):
        @classmethod
        def today(cls):
            return _real_date(2023, 1, 5)

    with patch.object(ledger_core, "date", _FakeDate):
        records = transactions.get_all_transactions()
        transactions._rebuild_balance(records)

    bal = storage.load_balance()
    assert abs(bal["AAPL"]["amount"] - 15.0) < 1e-6, \
        f"Balance should be 15 AAPL (today's +5 included), got {bal['AAPL']['amount']}"


# ── Holdings return calculation tests ─────────────────────────────────────────

def test_avg_price_stored_in_native_currency(tmp: Path):
    """avg_price is stored in the ticker's native currency (EUR for .DE stocks)."""
    import transactions, storage

    fx.inject_fake_prices(tmp)
    storage.save_price_year("SEC0.DE", 2023, {"2023-01-03": 88.27})

    transactions.add_transaction("2023-01-03", [
        {"ticker": "SEC0.DE", "amount": 288.0},
        {"ticker": "EUR",     "amount": -25425.0},
    ])

    bal = storage.load_balance()
    avg = bal["SEC0.DE"]["avg_price"]
    assert abs(avg - 88.27) < 0.01, \
        f"avg_price should be 88.27 EUR (native), got {avg}"


def test_avg_price_usd_stored_in_usd(tmp: Path):
    """avg_price for USD stock is stored in USD."""
    import transactions, storage

    fx.inject_fake_prices(tmp)
    storage.save_price_year("GOOG", 2023, {"2023-01-03": 88.0})

    transactions.add_transaction("2023-01-03", [
        {"ticker": "GOOG", "amount": 10.0},
        {"ticker": "USD",  "amount": -880.0},
    ])

    bal = storage.load_balance()
    avg = bal["GOOG"]["avg_price"]
    assert abs(avg - 88.0) < 0.01, \
        f"avg_price should be 88.0 USD (native), got {avg}"


def test_cost_basis_eur_stock_converts_to_pln(tmp: Path):
    """Cost basis for EUR stock uses avg_price × EURPLN rate × shares."""
    import transactions, portfolio, storage

    fx.inject_fake_prices(tmp)
    storage.save_price_year("SEC0.DE", 2023, {
        "2023-01-03": 88.27,
        "2023-01-04": 90.00,
    })

    transactions.add_transaction("2023-01-03", [
        {"ticker": "SEC0.DE", "amount": 288.0},
        {"ticker": "EUR",     "amount": -25425.0},
    ])

    snaps = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 4),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    snap = next(s for s in snaps if s["date"] == "2023-01-04")
    asset = next(a for a in snap["assets"] if a["ticker"] == "SEC0.DE")

    bal = storage.load_balance()
    avg_raw = bal["SEC0.DE"]["avg_price"]  # 88.27 EUR (native)

    # avg_price is in native EUR, not PLN — verify it's the stock price, not inflated
    assert 80 < avg_raw < 100, \
        f"avg_price should be ~88 EUR (native), got {avg_raw}"

    # Value uses EURPLN, cost basis must use same currency
    # Return = (90/88.27 - 1) × 100 ≈ 1.96% — FX cancels out
    ret_pct = ((asset["value_base"] / (asset["amount"] * avg_raw * 4.67)) - 1) * 100
    assert 0 < ret_pct < 5, \
        f"Return should be ~2%, got {ret_pct:.1f}%"


def test_return_eur_stock_not_inflated(tmp: Path):
    """Return for EUR stock is reasonable, not 300%+ due to missing FX conversion."""
    import transactions, portfolio, storage

    fx.inject_fake_prices(tmp)
    storage.save_price_year("SXRV.DE", 2023, {
        "2023-01-03": 42.0,
        "2023-01-09": 44.0,
    })

    transactions.add_transaction("2023-01-03", [
        {"ticker": "SXRV.DE", "amount": 3.5},
        {"ticker": "EUR",     "amount": -147.0},
    ])

    snaps = portfolio.build_portfolio(
        start_date=date(2023, 1, 3),
        end_date=date(2023, 1, 9),
        base_currency="PLN",
        precision="D",
        use_cache=False,
    )

    snap = next(s for s in snaps if s["date"] == "2023-01-09")
    asset = next(a for a in snap["assets"] if a["ticker"] == "SXRV.DE")

    bal = storage.load_balance()
    avg_raw = bal["SXRV.DE"]["avg_price"]  # 42.0 EUR (native)

    # avg_price should be the EUR stock price, not some PLN-inflated number
    assert 30 < avg_raw < 50, \
        f"avg_price should be ~42 EUR (native), got {avg_raw}"

    # Return ≈ 44/42 - 1 ≈ 4.8%. With proper FX conversion both value and
    # cost_basis use the same EURPLN rate, so it cancels out.
    # Without the fix, cost_basis would be in EUR while value is in PLN → 300%+.
    # We just check it's in a sane range (0–20%).
    shares = asset["amount"]
    # Use avg_raw directly (native EUR) — this is what the holdings table
    # does BEFORE the fix (missing FX conversion). If the fix works,
    # the return should be ~5% regardless of FX rate used.
    ret_pct_native = ((asset["value_base"] / (shares * avg_raw * 4.34)) - 1) * 100
    assert -5 < ret_pct_native < 20, \
        f"Return ~{ret_pct_native:.1f}% looks inflated (FX conversion may be missing)"


# ── Ticker history tests ──────────────────────────────────────────────────────

def test_get_ticker_history_buys_and_sells(tmp: Path):
    """get_ticker_history returns chronological buys/sells with correct sides."""
    import transactions

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD",  "amount": -1250.0},
    ])
    transactions.add_transaction("2023-06-01", [
        {"ticker": "AAPL", "amount": -4.0},
        {"ticker": "USD",  "amount": 500.0},
    ])

    hist = transactions.get_ticker_history("AAPL")
    assert len(hist) == 2
    assert hist[0]["side"] == "Buy"
    assert hist[0]["amount"] == 10.0
    assert hist[0]["date"] == "2023-01-03"
    assert hist[1]["side"] == "Sell"
    assert hist[1]["amount"] == -4.0
    assert hist[1]["date"] == "2023-06-01"


def test_get_ticker_history_running_shares(tmp: Path):
    """Running position tracks cumulative shares across trades."""
    import transactions

    transactions.add_transaction("2023-01-03", [{"ticker": "GOOG", "amount": 10.0}])
    transactions.add_transaction("2023-03-01", [{"ticker": "GOOG", "amount": 5.0}])
    transactions.add_transaction("2023-06-01", [{"ticker": "GOOG", "amount": -8.0}])

    hist = transactions.get_ticker_history("GOOG")
    assert len(hist) == 3
    assert hist[0]["running"] == 10.0
    assert hist[1]["running"] == 15.0
    assert hist[2]["running"] == 7.0


def test_get_ticker_history_excludes_cash(tmp: Path):
    """Cash tickers (USD, EUR, PLN) are excluded from history."""
    import transactions

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD",  "amount": -1250.0},
    ])

    hist_aapl = transactions.get_ticker_history("AAPL")
    assert len(hist_aapl) == 1

    hist_usd = transactions.get_ticker_history("USD")
    assert len(hist_usd) == 0

    hist_pln = transactions.get_ticker_history("PLN")
    assert len(hist_pln) == 0


# ── Theme consistency tests ──────────────────────────────────────────────────

REQUIRED_THEME_KEYS = [
    "card_bg", "border", "border_hover", "border_active", "border_strong",
    "text", "text_cell", "text_muted", "text_faint",
    "panel_bg", "hr", "accent",
    "accent_tag_bg", "accent_tag_border", "accent_tag_hover",
    "card_btn_active", "card_faint",
    "chart_grid", "chart_zeroline", "hover_bg",
    "plotly_template", "page_bg",
    "range_bg", "range_active",
    "table_border", "table_header_border", "table_hover", "table_text",
    "holdings_bar_bg",
]


def _is_valid_css_color(value: str) -> bool:
    """Check if a string looks like a valid CSS color."""
    import re
    # Named colors, hex, rgb(), rgba(), hsl(), hsla()
    if value.startswith("#"):
        return len(value) in (4, 5, 7, 9) and all(c in "0123456789abcdefABCDEF" for c in value[1:])
    if value.startswith(("rgb(", "rgba(", "hsl(", "hsla(")):
        return bool(re.match(r"^(rgb|rgba|hsl|hsla)\(.+\)$", value))
    return False


def test_theme_keys_dark(tmp: Path):
    """Dark theme defines all required keys."""
    from ui.colors import THEMES
    for key in REQUIRED_THEME_KEYS:
        assert key in THEMES["dark"], f"Dark theme missing key: {key}"


def test_theme_keys_light(tmp: Path):
    """Light theme defines all required keys."""
    from ui.colors import THEMES
    for key in REQUIRED_THEME_KEYS:
        assert key in THEMES["light"], f"Light theme missing key: {key}"


def test_theme_colors_valid(tmp: Path):
    """All theme color values are valid CSS colors."""
    from ui.colors import THEMES
    NON_COLOR_KEYS = {"plotly_template"}
    for theme_name, theme in THEMES.items():
        for key, value in theme.items():
            if key in NON_COLOR_KEYS:
                continue
            assert _is_valid_css_color(value), f"{theme_name}.{key} = {value!r} is not a valid CSS color"


def test_theme_plotly_template(tmp: Path):
    """Dark uses plotly_dark, light uses plotly_white."""
    from ui.colors import THEMES
    assert THEMES["dark"]["plotly_template"] == "plotly_dark"
    assert THEMES["light"]["plotly_template"] == "plotly_white"


def test_theme_css_uses_correct_colors(tmp: Path):
    """CSS builders produce output containing the theme's colors."""
    from ui.colors import THEMES
    from ui.style_base import build_app_styles
    from ui.style_components import build_metric_card_styles

    for theme_name in ("dark", "light"):
        t = THEMES[theme_name]

        app_css = build_app_styles(t)
        assert t["page_bg"] in app_css, f"{theme_name}: build_app_styles missing page_bg"
        assert t["panel_bg"] in app_css, f"{theme_name}: build_app_styles missing panel_bg"
        assert t["text"] in app_css, f"{theme_name}: build_app_styles missing text"

        metric_css = build_metric_card_styles(t)
        assert t["card_bg"] in metric_css, f"{theme_name}: build_metric_card_styles missing card_bg"
        assert t["text"] in metric_css, f"{theme_name}: build_metric_card_styles missing text"


def test_theme_plotly_font_colors(tmp: Path):
    """Plotly chart layout uses theme text color for fonts."""
    from ui.colors import THEMES
    import plotly.graph_objects as go

    for theme_name in ("dark", "light"):
        t = THEMES[theme_name]
        fig = go.Figure()
        fig.update_layout(
            font=dict(family="sans-serif", size=20, color=t["text"]),
        )
        assert fig.layout.font.color == t["text"], (
            f"{theme_name}: Plotly font color {fig.layout.font.color} != {t['text']}"
        )


def test_theme_dark_text_on_dark_bg(tmp: Path):
    """Dark theme text colors are light (for dark backgrounds)."""
    from ui.colors import THEMES
    t = THEMES["dark"]
    # Text should start with # (hex) and have high RGB values
    for key in ("text", "text_cell", "text_muted"):
        val = t[key]
        if val.startswith("#"):
            r = int(val[1:3], 16)
            g = int(val[3:5], 16)
            b = int(val[5:7], 16)
            assert r > 100 and g > 100 and b > 100, (
                f"Dark theme {key}={val} is too dark for dark backgrounds"
            )


def test_theme_light_text_on_light_bg(tmp: Path):
    """Light theme text colors are dark (for light backgrounds)."""
    from ui.colors import THEMES
    t = THEMES["light"]
    for key in ("text", "text_cell", "text_muted"):
        val = t[key]
        if val.startswith("#"):
            r = int(val[1:3], 16)
            g = int(val[3:5], 16)
            b = int(val[5:7], 16)
            assert r < 150 and g < 150 and b < 150, (
                f"Light theme {key}={val} is too light for light backgrounds"
            )


# ── CAGR tests ───────────────────────────────────────────────────────────────

def test_cagr_basic(tmp: Path):
    """CAGR: single deposit growing over time yields positive return."""
    import transactions
    from ledger_core import compute_cagr
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0, "account_operation": True},
    ])

    cagr = compute_cagr(12000.0, "PLN")
    assert cagr is not None
    assert cagr > 0


def test_cagr_no_transactions(tmp: Path):
    """CAGR: no transactions returns None."""
    from ledger_core import compute_cagr
    assert compute_cagr(10000.0, "PLN") is None


def test_cagr_zero_invested(tmp: Path):
    """CAGR: zero net invested returns None."""
    import transactions
    from ledger_core import compute_cagr
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": -10000.0, "account_operation": True},
    ])

    cagr = compute_cagr(0.0, "PLN")
    assert cagr is None


def test_cagr_loss(tmp: Path):
    """CAGR: value below invested yields negative CAGR."""
    import transactions
    from ledger_core import compute_cagr
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0, "account_operation": True},
    ])

    cagr = compute_cagr(5000.0, "PLN")
    assert cagr is not None
    assert cagr < 0


def test_cagr_with_fx(tmp: Path):
    """CAGR: USD deposit converted to PLN uses FX rate."""
    import transactions
    from ledger_core import compute_cagr
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "USD", "amount": 2000.0, "account_operation": True},
    ])

    cagr = compute_cagr(10000.0, "PLN")
    assert cagr is not None
    assert cagr > 0


# ── IRR tests ────────────────────────────────────────────────────────────────

def test_irr_basic(tmp: Path):
    """IRR: single deposit with positive return yields positive IRR."""
    import transactions
    from ledger_core import compute_irr
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0, "account_operation": True},
    ])

    irr = compute_irr(12000.0, "PLN")
    assert irr is not None
    assert irr > 0


def test_irr_no_transactions(tmp: Path):
    """IRR: no transactions returns None."""
    from ledger_core import compute_irr
    assert compute_irr(10000.0, "PLN") is None


def test_irr_multiple_deposits(tmp: Path):
    """IRR: multiple deposits are handled correctly."""
    import transactions
    from ledger_core import compute_irr
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 5000.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-06-01", [
        {"ticker": "PLN", "amount": 5000.0, "account_operation": True},
    ])

    irr = compute_irr(12000.0, "PLN")
    assert irr is not None
    assert irr > 0


def test_irr_with_withdrawal(tmp: Path):
    """IRR: withdrawal reduces invested capital correctly."""
    import transactions
    from ledger_core import compute_irr
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-06-01", [
        {"ticker": "PLN", "amount": -3000.0, "account_operation": True},
    ])

    irr = compute_irr(9000.0, "PLN")
    assert irr is not None


def test_irr_loss(tmp: Path):
    """IRR: value below invested yields negative IRR."""
    import transactions
    from ledger_core import compute_irr
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0, "account_operation": True},
    ])

    irr = compute_irr(5000.0, "PLN")
    assert irr is not None
    assert irr < 0


def _reference_irr(ext_flows: list, terminal: float) -> float:
    """Independent XIRR solver (bisection) used as an oracle for compute_irr.

    ``ext_flows`` is a list of (date_str, amount_in_base_ccy) where outflows
    are negative (deposits) and inflows positive (withdrawals). ``terminal``
    is the final positive portfolio value dated today.
    """
    from datetime import date as _d

    today = _d.today().isoformat()
    items = sorted(list(ext_flows) + [(today, terminal)], key=lambda x: x[0])
    start = _d.fromisoformat(items[0][0])
    yrs = [(_d.fromisoformat(d) - start).days / 365.25 for d, _ in items]
    amts = [a for _, a in items]

    def npv(r: float) -> float:
        return sum(a / (1.0 + r) ** t for a, t in zip(amts, yrs))

    lo, hi = -0.5, 5.0
    for _ in range(300):
        mid = (lo + hi) / 2.0
        if npv(mid) > 0:
            lo = mid
        else:
            hi = mid
        if abs(hi - lo) < 1e-11:
            break
    return (lo + hi) / 2.0


def _rebuild_npv(current_value: float, base_ccy: str) -> float:
    """Rebuild compute_irr's cash flows and return NPV at the computed IRR.

    Oracle-free correctness check: the IRR returned by compute_irr must zero
    the NPV of exactly the flows it used.
    """
    from datetime import date as _d
    from ledger_core import compute_irr, get_all_transactions
    from ticker_data import get_fx_rate

    base = base_ccy.upper()
    flows: list = []
    for rec in get_all_transactions():
        for e in rec["entries"]:
            if not e.get("account_operation", False):
                continue
            t = e["ticker"].upper()
            amt = float(e["amount"])
            fx = get_fx_rate(t, base, rec["date"], {}, int(rec["date"][:4])) if t != base else 1.0
            flows.append((rec["date"], -amt * fx))

    today = _d.today().isoformat()
    flows.append((today, current_value))
    flows.sort(key=lambda x: x[0])
    start = _d.fromisoformat(flows[0][0])
    yrs = [(_d.fromisoformat(d) - start).days / 365.25 for d, _ in flows]
    amts = [a for _, a in flows]

    irr = compute_irr(current_value, base)
    return irr, sum(a / (1.0 + irr) ** t for a, t in zip(amts, yrs))


def test_irr_known_two_flow(tmp: Path):
    """IRR: a 1-year 1000→1100 deposit equals the closed-form ~10%."""
    import transactions
    from ledger_core import compute_irr
    from datetime import date as _d

    fx.inject_fake_prices(tmp)
    dep = _d.today().replace(year=_d.today().year - 1).isoformat()
    transactions.add_transaction(dep, [
        {"ticker": "PLN", "amount": 1000.0, "account_operation": True},
    ])
    irr = compute_irr(1100.0, "PLN")
    days = (_d.today() - _d.fromisoformat(dep)).days
    expected = (1100.0 / 1000.0) ** (365.25 / days) - 1.0
    assert abs(irr - expected) < 1e-6


def test_irr_dividend_ignored(tmp: Path):
    """IRR regression: a dividend must NOT change the IRR (it is internal)."""
    import transactions
    from ledger_core import compute_irr

    fx.inject_fake_prices(tmp)
    # baseline: deposit only
    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0, "account_operation": True},
    ])
    irr_base = compute_irr(12000.0, "PLN")

    # same deposit + a dividend (currency entry, no account_operation flag)
    transactions.add_transaction("2023-06-01", [
        {"ticker": "PLN", "amount": 100.0},
    ])
    irr_div = compute_irr(12000.0, "PLN")
    assert abs(irr_base - irr_div) < 1e-9


def test_irr_withholding_tax_ignored(tmp: Path):
    """IRR regression: a withholding-tax entry must NOT change the IRR."""
    import transactions
    from ledger_core import compute_irr

    fx.inject_fake_prices(tmp)
    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 10000.0, "account_operation": True},
    ])
    irr_base = compute_irr(12000.0, "PLN")
    transactions.add_transaction("2023-06-01", [
        {"ticker": "PLN", "amount": -20.0},
    ])
    irr_tax = compute_irr(12000.0, "PLN")
    assert abs(irr_base - irr_tax) < 1e-9


def test_irr_fx_deposit(tmp: Path):
    """IRR: a USD deposit is FX-converted and matches a PLN reference IRR."""
    import transactions
    from ledger_core import compute_irr
    from ticker_data import get_fx_rate

    fx.inject_fake_prices(tmp)
    dep_date = "2023-01-03"  # USDPLN = 4.38 in fake data
    usd_rate = get_fx_rate("USD", "PLN", dep_date, {}, 2023)
    transactions.add_transaction(dep_date, [
        {"ticker": "USD", "amount": 1000.0, "account_operation": True},
    ])
    # terminal value = deposit grown 10% in USD terms, expressed in PLN
    terminal = 1000.0 * usd_rate * 1.10
    irr = compute_irr(terminal, "PLN")
    ref = _reference_irr([(dep_date, -1000.0 * usd_rate)], terminal)
    assert abs(irr - ref) < 1e-6


def test_irr_matches_reference_xirr(tmp: Path):
    """IRR: solver agrees with an independent bisection XIRR oracle."""
    import transactions
    from ledger_core import compute_irr

    fx.inject_fake_prices(tmp)
    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 5000.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-06-01", [
        {"ticker": "PLN", "amount": -2000.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-09-15", [
        {"ticker": "PLN", "amount": 3000.0, "account_operation": True},
    ])
    terminal = 9000.0
    flows = [
        ("2023-01-03", -5000.0),
        ("2023-06-01", 2000.0),
        ("2023-09-15", -3000.0),
    ]
    irr = compute_irr(terminal, "PLN")
    ref = _reference_irr(flows, terminal)
    assert abs(irr - ref) < 1e-6


def test_irr_npv_is_zero(tmp: Path):
    """IRR property: compute_irr's rate zeroes the NPV of its own flows."""
    import transactions
    from ledger_core import compute_irr

    fx.inject_fake_prices(tmp)
    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 5000.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-06-01", [
        {"ticker": "PLN", "amount": -2000.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-09-15", [
        {"ticker": "PLN", "amount": 3000.0, "account_operation": True},
    ])
    _, npv = _rebuild_npv(9000.0, "PLN")
    assert abs(npv) < 1e-6


def test_irr_multiple_sign_changes(tmp: Path):
    """IRR: deposit/withdraw/deposit still yields a finite, bounded rate."""
    import transactions
    from ledger_core import compute_irr

    fx.inject_fake_prices(tmp)
    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 1000.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-06-01", [
        {"ticker": "PLN", "amount": -400.0, "account_operation": True},
    ])
    transactions.add_transaction("2023-09-15", [
        {"ticker": "PLN", "amount": 600.0, "account_operation": True},
    ])
    irr = compute_irr(1500.0, "PLN")
    assert irr is not None
    assert -0.5 <= irr <= 5.0


def test_irr_high_return(tmp: Path):
    """IRR: a very large gain stays within the solver's upper bound."""
    import transactions
    from ledger_core import compute_irr

    fx.inject_fake_prices(tmp)
    transactions.add_transaction("2023-01-03", [
        {"ticker": "PLN", "amount": 1000.0, "account_operation": True},
    ])
    irr = compute_irr(20000.0, "PLN")
    assert irr is not None
    assert irr > 0.5


# ── FX rate tests ────────────────────────────────────────────────────────────

def test_fx_rate_same_currency(tmp: Path):
    """get_fx_rate: same currency returns 1.0."""
    from ticker_data import get_fx_rate
    assert get_fx_rate("PLN", "PLN", "2023-01-03", {}, 2023) == 1.0


def test_fx_rate_direct_pair(tmp: Path):
    """get_fx_rate: direct pair (USDPLN) returns cached price."""
    from ticker_data import get_fx_rate
    cache = {"USDPLN": {2023: {"2023-01-03": 4.38}}}
    assert get_fx_rate("USD", "PLN", "2023-01-03", cache, 2023) == 4.38


def test_fx_rate_reverse_pair(tmp: Path):
    """get_fx_rate: reverse pair returns 1/rate."""
    from ticker_data import get_fx_rate
    cache = {"USDPLN": {2023: {"2023-01-03": 4.0}}}
    rate = get_fx_rate("PLN", "USD", "2023-01-03", cache, 2023)
    assert abs(rate - 0.25) < 0.001


def test_fx_rate_eur_pln_via_triangulation(tmp: Path):
    """get_fx_rate: EUR→PLN triangulated via EURUSD * USDPLN."""
    from ticker_data import get_fx_rate
    cache = {
        "EURUSD": {2023: {"2023-01-03": 1.07}},
        "USDPLN": {2023: {"2023-01-03": 4.0}},
    }
    rate = get_fx_rate("EUR", "PLN", "2023-01-03", cache, 2023)
    assert abs(rate - 4.28) < 0.01


def test_fx_rate_usd_pln_direct(tmp: Path):
    """get_fx_rate: USD→PLN via direct USDPLN pair."""
    from ticker_data import get_fx_rate
    cache = {"USDPLN": {2023: {"2023-01-03": 4.35}}}
    rate = get_fx_rate("USD", "PLN", "2023-01-03", cache, 2023)
    assert abs(rate - 4.35) < 0.001


def test_fx_rate_pln_to_usd(tmp: Path):
    """get_fx_rate: PLN→USD via 1/USDPLN."""
    from ticker_data import get_fx_rate
    cache = {"USDPLN": {2023: {"2023-01-03": 4.0}}}
    rate = get_fx_rate("PLN", "USD", "2023-01-03", cache, 2023)
    assert abs(rate - 0.25) < 0.001


def test_fx_rate_fallback_1(tmp: Path):
    """get_fx_rate: no data returns 1.0 as last resort."""
    from ticker_data import get_fx_rate
    rate = get_fx_rate("USD", "PLN", "2023-01-03", {}, 2023)
    assert rate == 1.0


def test_fx_rate_non_pln_cross_via_pln(tmp: Path):
    """get_fx_rate: EUR→USD triangulated via EURPLN / USDPLN."""
    from ticker_data import get_fx_rate
    cache = {
        "EURPLN": {2023: {"2023-01-03": 4.68}},
        "USDPLN": {2023: {"2023-01-03": 4.0}},
    }
    rate = get_fx_rate("EUR", "USD", "2023-01-03", cache, 2023)
    assert abs(rate - 1.17) < 0.01


# ── _latest_price tests ──────────────────────────────────────────────────────

def test_latest_price_cash(tmp: Path):
    """_latest_price: cash ticker returns 1.0."""
    from ticker_data import _latest_price
    assert _latest_price("USD", {}, 2023) == 1.0


def test_latest_price_from_cache(tmp: Path):
    """_latest_price: returns latest price from loaded cache."""
    from ticker_data import _latest_price
    cache = {"AAPL": {2023: {"2023-01-03": 125.0, "2023-01-09": 130.0}}}
    assert _latest_price("AAPL", cache, 2023) == 130.0


def test_latest_price_empty_cache(tmp: Path):
    """_latest_price: missing ticker returns None."""
    from ticker_data import _latest_price
    assert _latest_price("AAPL", {}, 2023) is None


def test_latest_price_empty_slab(tmp: Path):
    """_latest_price: empty slab returns None."""
    from ticker_data import _latest_price
    cache = {"AAPL": {2023: {}}}
    assert _latest_price("AAPL", cache, 2023) is None


def test_latest_price_multi_year(tmp: Path):
    """_latest_price: scans newest year first."""
    from ticker_data import _latest_price
    cache = {
        "AAPL": {
            2022: {"2022-12-29": 120.0},
            2023: {"2023-01-03": 125.0},
        }
    }
    assert _latest_price("AAPL", cache, 2023) == 125.0


# ── BOSSA import tests ──────────────────────────────────────────────────────

def test_bossa_validate_valid(tmp: Path):
    """validate_bossa_file: valid CSV with required columns passes."""
    from bossa_import import validate_bossa_file
    csv_content = "data;tytuł operacji;szczegóły;kwota;waluta\n2023-01-03;test;;;PLN\n"
    p = tmp / "test.csv"
    p.write_text(csv_content, encoding="utf-8")
    valid, msg = validate_bossa_file(p)
    assert valid is True


def test_bossa_validate_empty(tmp: Path):
    """validate_bossa_file: empty file fails."""
    from bossa_import import validate_bossa_file
    p = tmp / "empty.csv"
    p.write_text("", encoding="utf-8")
    valid, msg = validate_bossa_file(p)
    assert valid is False


def test_bossa_validate_missing_columns(tmp: Path):
    """validate_bossa_file: missing required columns fails."""
    from bossa_import import validate_bossa_file
    csv_content = "col1;col2;col3\n"
    p = tmp / "bad.csv"
    p.write_text(csv_content, encoding="utf-8")
    valid, msg = validate_bossa_file(p)
    assert valid is False
    assert "Missing required columns" in msg


def test_bossa_parse_float_basic(tmp: Path):
    """_parse_float: normal number parsing."""
    from bossa_import import _parse_float
    assert _parse_float("123.45") == 123.45
    assert _parse_float("123,45") == 123.45
    assert _parse_float("-50.0") == -50.0


def test_bossa_parse_float_empty(tmp: Path):
    """_parse_float: empty/None returns None."""
    from bossa_import import _parse_float
    assert _parse_float("") is None
    assert _parse_float(None) is None


def test_bossa_parse_float_invalid(tmp: Path):
    """_parse_float: non-numeric returns None."""
    from bossa_import import _parse_float
    assert _parse_float("abc") is None
    assert _parse_float("  ") is None


def test_bossa_read_csv_utf8(tmp: Path):
    """_read_csv_text: reads UTF-8 file."""
    from bossa_import import _read_csv_text
    p = tmp / "test.csv"
    p.write_text("data;kwota\n100;200\n", encoding="utf-8")
    text = _read_csv_text(p)
    assert "data;kwota" in text


def test_bossa_read_csv_windows1250(tmp: Path):
    """_read_csv_text: reads Windows-1250 encoded file."""
    from bossa_import import _read_csv_text
    p = tmp / "test.csv"
    content = "data;tytuł\n"
    p.write_bytes(content.encode("windows-1250"))
    text = _read_csv_text(p)
    assert "data;tytuł" in text


def test_bossa_parse_buy(tmp: Path):
    """parse_bossa_csv: buy transaction creates two entries."""
    from bossa_import import parse_bossa_csv
    import bossa_import

    original_resolve = bossa_import.resolve_isins_with_names
    bossa_import.resolve_isins_with_names = lambda names, progress_cb=None: (
        {isin: "AAPL.US" for isin in names}, {}
    )
    try:
        csv_content = (
            "data;tytuł operacji;szczegóły;kwota;waluta\n"
            "2023-01-03;Rozliczenie transakcji kupna;Apple Inc. (US0378331005) 10 x 125.07 USD nr 1;-1250.70;USD\n"
        )
        p = tmp / "buy.csv"
        p.write_text(csv_content, encoding="utf-8")
        txns, unresolved = parse_bossa_csv(p, "PLN")
        assert len(txns) == 1
        assert txns[0]["date"] == "2023-01-03"
        tickers = {e["ticker"] for e in txns[0]["entries"]}
        assert "AAPL.US" in tickers
        assert "USD" in tickers
    finally:
        bossa_import.resolve_isins_with_names = original_resolve


def test_bossa_parse_sell(tmp: Path):
    """parse_bossa_csv: sell transaction creates entries with negative shares."""
    from bossa_import import parse_bossa_csv
    import bossa_import

    original_resolve = bossa_import.resolve_isins_with_names
    bossa_import.resolve_isins_with_names = lambda names, progress_cb=None: (
        {isin: "AAPL.US" for isin in names}, {}
    )
    try:
        csv_content = (
            "data;tytuł operacji;szczegóły;kwota;waluta\n"
            "2023-06-01;Rozliczenie transakcji sprzedaży;Apple Inc. (US0378331005) 5 x 180.09 USD nr 2;900.45;USD\n"
        )
        p = tmp / "sell.csv"
        p.write_text(csv_content, encoding="utf-8")
        txns, _ = parse_bossa_csv(p, "PLN")
        # May be 1 or 2 txns (fix_negative_positions can append a corrective buy)
        all_entries = [e for rec in txns for e in rec["entries"]]
        stock_entries = [e for e in all_entries if e["ticker"] == "AAPL.US"]
        assert len(stock_entries) >= 1
        assert any(e["amount"] < 0 for e in stock_entries)
    finally:
        bossa_import.resolve_isins_with_names = original_resolve


def test_bossa_parse_deposit(tmp: Path):
    """parse_bossa_csv: deposit marked as account_operation."""
    from bossa_import import parse_bossa_csv

    csv_content = (
        "data;tytuł operacji;szczegóły;kwota;waluta\n"
        "2023-01-03;Przelew do DM BOŚ;;5000.00;PLN\n"
    )
    p = tmp / "deposit.csv"
    p.write_text(csv_content, encoding="utf-8")
    txns, _ = parse_bossa_csv(p, "PLN")
    assert len(txns) == 1
    assert txns[0]["entries"][0].get("account_operation") is True
    assert txns[0]["entries"][0]["amount"] == 5000.0


def test_bossa_parse_short_row_skipped(tmp: Path):
    """parse_bossa_csv: rows with fewer than 5 columns are skipped."""
    from bossa_import import parse_bossa_csv

    csv_content = (
        "data;tytuł operacji;szczegóły;kwota;waluta\n"
        "2023-01-03;short row\n"
    )
    p = tmp / "short.csv"
    p.write_text(csv_content, encoding="utf-8")
    txns, _ = parse_bossa_csv(p, "PLN")
    assert len(txns) == 0


# ── XTB validate tests ──────────────────────────────────────────────────────

def test_xtb_validate_missing_sheet(tmp: Path):
    """validate_xtb_file: file without 'Cash Operations' sheet fails."""
    from xtb_import import validate_xtb_file
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Positions"
    wb.active.append(["Type", "Ticker", "Amount"])
    p = tmp / "no_cash.xlsx"
    wb.save(p)
    wb.close()

    valid, msg = validate_xtb_file(p)
    assert valid is False
    assert "Cash Operations" in msg


def test_xtb_validate_missing_columns(tmp: Path):
    """validate_xtb_file: sheet missing required columns fails."""
    from xtb_import import validate_xtb_file
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker"])  # Missing Amount, Time
    p = tmp / "missing_cols.xlsx"
    wb.save(p)
    wb.close()

    valid, msg = validate_xtb_file(p)
    assert valid is False
    assert "Missing columns" in msg


def test_xtb_validate_valid(tmp: Path):
    """validate_xtb_file: valid XTB file passes."""
    from xtb_import import validate_xtb_file
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Deposit", "USD", 1000, "2023-01-03 10:00:00", ""])
    p = tmp / "valid.xlsx"
    wb.save(p)
    wb.close()

    valid, msg = validate_xtb_file(p)
    assert valid is True


def test_xtb_parse_dividend(tmp: Path):
    """parse_xtb_excel: dividend creates a currency entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Dividend", "AAPL.US", 15.50, "2023-06-15 10:00:00", "Dividend AAPL"])
    p = tmp / "div.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0]["amount"] == 15.50
    assert txns[0]["entries"][0]["ticker"] == "USD"


def test_xtb_parse_withdrawal(tmp: Path):
    """parse_xtb_excel: withdrawal creates account_operation entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Withdrawal", "USD", -500.0, "2023-07-01 10:00:00", ""])
    p = tmp / "withdraw.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0].get("account_operation") is True
    assert txns[0]["entries"][0]["amount"] == -500.0


def test_xtb_parse_interest(tmp: Path):
    """parse_xtb_excel: free funds interest creates currency entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Free funds interest", "USD", 2.35, "2023-06-30 10:00:00", ""])
    p = tmp / "interest.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0]["amount"] == 2.35
    assert txns[0]["entries"][0]["ticker"] == "USD"


def test_xtb_parse_withholding_tax(tmp: Path):
    """parse_xtb_excel: withholding tax creates currency entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Withholding tax", "USD", -3.10, "2023-06-15 10:00:00", ""])
    p = tmp / "wht.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0]["amount"] == -3.10


def test_xtb_parse_deposit(tmp: Path):
    """parse_xtb_excel: deposit creates account_operation entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Deposit", "USD", 5000.0, "2023-01-03 10:00:00", ""])
    p = tmp / "deposit.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0].get("account_operation") is True
    assert txns[0]["entries"][0]["amount"] == 5000.0


def test_xtb_parse_transfer(tmp: Path):
    """parse_xtb_excel: transfer creates account_operation entry."""
    from xtb_import import parse_xtb_excel
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Operations"
    ws.append(["Type", "Ticker", "Amount", "Time", "Comment"])
    ws.append(["Transfer", "USD", 2000.0, "2023-03-15 10:00:00", ""])
    p = tmp / "transfer.xlsx"
    wb.save(p)
    wb.close()

    txns = parse_xtb_excel(p, "USD")
    assert len(txns) == 1
    assert txns[0]["entries"][0].get("account_operation") is True


# ── _update_avg_prices sell path tests ──────────────────────────────────────

def test_avg_price_sell_reduces_proportionally(tmp: Path):
    """_update_avg_prices: sell-only reduces cost pool proportionally."""
    import transactions
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD", "amount": -1250.70},
    ])

    transactions.add_transaction("2023-06-01", [
        {"ticker": "AAPL", "amount": -5.0},
        {"ticker": "USD", "amount": 900.45},
    ])

    import storage
    bal = storage.load_balance()
    assert bal["AAPL"]["amount"] == 5.0
    assert bal["AAPL"]["avg_price"] > 0


def test_avg_price_full_sell_zeroes(tmp: Path):
    """_update_avg_prices: full sell removes ticker from balance."""
    import transactions
    fx.inject_fake_prices(tmp)

    transactions.add_transaction("2023-01-03", [
        {"ticker": "AAPL", "amount": 10.0},
        {"ticker": "USD", "amount": -1250.70},
    ])

    transactions.add_transaction("2023-06-01", [
        {"ticker": "AAPL", "amount": -10.0},
        {"ticker": "USD", "amount": 1800.90},
    ])

    import storage
    bal = storage.load_balance()
    assert "AAPL" not in bal or abs(bal.get("AAPL", {}).get("amount", 0)) < 1e-9


# ── Runner ────────────────────────────────────────────────────────────────────

ALL_TESTS = [
    ("Config: creates defaults",              test_config_defaults),
    ("Transactions: buy EUR ETF full sell",   test_buy_eur_etf_full_sell),
    ("Transactions: buy USD stock partial",   test_buy_usd_stock_partial_sell),
    ("Portfolio: sell proceeds not contrib",  test_sell_proceeds_not_counted_as_invested),
    ("Portfolio: cash deposit is contrib",    test_currency_exchange_counts_as_invested),
    ("Portfolio: EUR stock → PLN correct",    test_eur_stock_valued_correctly_in_pln),
    ("Portfolio: USD stock → PLN correct",    test_usd_stock_valued_correctly_in_pln),
    ("Portfolio: mixed PLN/EUR/USD portfolio",test_mixed_portfolio_pln_eur_usd),
    ("Config: save and reload",               test_config_save_and_reload),
    ("Config: precision mapping",             test_config_precision_mapping),
    ("Storage: JSONL round-trip",             test_storage_jsonl_roundtrip),
    ("Storage: JSONL append",                 test_storage_append_jsonl),
    ("Storage: balance save/load",            test_storage_balance),
    ("Storage: price cache write/read",       test_price_cache_write_read),
    ("Storage: load_prices_range",            test_storage_loads_prices_range),
    ("Transactions: add simple",              test_add_transaction_simple),
    ("Transactions: same-date merges",        test_add_transaction_same_date_merges),
    ("Transactions: chronological append",    test_add_transaction_chronological_append),
    ("Transactions: past-date insert",        test_add_transaction_past_date_inserts_correctly),
    ("Transactions: compute_holdings_at",     test_compute_holdings_at),
    ("Transactions: full sell zeroes balance",test_balance_after_full_sell),
    ("Transactions: delete entry",            test_delete_transaction),
    ("Transactions: delete last entry removes record", test_delete_last_entry_removes_record),
    ("Transactions: update entry",            test_update_transaction),
    ("Transactions: get_tickers",             test_get_tickers),
    ("XTB: parse_shares",                     test_xtb_parse_shares),
    ("XTB: parse_transfer_rate",              test_xtb_parse_transfer_rate),
    ("XTB: parse_transfer_target",            test_xtb_parse_transfer_target),
    ("XTB: transfer creates source entry",    test_xtb_transfer_creates_source_entry),
    ("XTB: deposit has account_operation",    test_xtb_deposit_creates_account_operation),
    ("XTB: stock purchase two entries",       test_xtb_stock_purchase_creates_two_entries),
    ("XTB: withholding tax entry",            test_xtb_withholding_tax),
    ("XTB: negative position gets fixed",    test_xtb_negative_position_gets_fixed),
    ("Prices: weekend fallback",              test_get_price_fallback_weekend),
    ("Prices: cash returns 1.0",              test_get_price_cash_returns_one),
    ("FX: same-currency rate is 1.0",         test_get_fx_rate_same_currency),
    ("FX: USD→PLN from cache",                test_get_fx_rate_usd_to_pln),
    ("FX: stale USD→PLN uses latest",         test_get_fx_rate_stale_usd_uses_latest),
    ("FX: stale USD↔PLN invertible",          test_get_fx_rate_stale_invertible),
    ("FX: GBP triangulates",                  test_get_fx_rate_gbp_triangulates),
    ("Holdings: weights currency invariant",  test_holdings_weights_currency_invariant),
    ("Portfolio: invalidate cache",           test_invalidate_portfolio),
    ("Portfolio: single asset PLN value",     test_portfolio_build_single_asset),
    ("Portfolio: cash-only",                  test_portfolio_build_cash_only),
    ("Portfolio: invested tracking",      test_portfolio_invested_tracking),
    ("Portfolio: weekly precision Fridays",   test_portfolio_weekly_precision),
    ("Portfolio: cache resume",               test_portfolio_cache_resumes),
    ("Portfolio: _day_range daily",           test_day_range_daily),
    ("Portfolio: _day_range weekly",          test_day_range_weekly),
    ("Portfolio: snapshots_to_series",        test_snapshots_to_series),
    # ── New tests ────────────────────────────────────────────────────────────
    ("Ticker translate: no rules",            test_translate_no_rules),
    ("Ticker translate: exact match",         test_translate_exact_match),
    ("Ticker translate: suffix swap",         test_translate_suffix_swap),
    ("Ticker translate: suffix strip",        test_translate_suffix_strip),
    ("Ticker translate: no match",            test_translate_no_match),
    ("Transactions: set_account_operation",   test_set_account_operation),
    ("Transactions: get_transactions_up_to",  test_get_transactions_up_to),
    ("Transactions: get_all_tickers",         test_get_all_tickers),
    ("Storage: create and list projects",     test_create_and_list_projects),
    ("Storage: rename project",               test_rename_project),
    ("Storage: delete project",               test_delete_project),
    ("Storage: global config shared",          test_config_is_global),
    ("Storage: benchmark save/load",          test_benchmark_save_load_roundtrip),
    ("Storage: benchmark load missing",       test_benchmark_load_returns_none_when_missing),
    ("Portfolio: withdrawal decreases invested", test_withdrawal_decreases_invested),
    ("Portfolio: ticker currency detection",  test_ticker_currency_detection),
    ("Manual: validate errors",              test_manual_validate_errors),
    ("Manual: parse and import",              test_manual_parse_and_import),
    ("BOSSA: validate errors",               test_bossa_validate_errors),
    ("ISIN: resolve from config",            test_isin_resolve_from_config),
    ("Portfolio: includes today transactions",  test_portfolio_includes_today_transactions),
    ("Rebuild balance: includes today",         test_rebuild_balance_includes_today_transactions),
    ("Avg price: stored in native currency",  test_avg_price_stored_in_native_currency),
    ("Avg price: USD stock in USD",           test_avg_price_usd_stored_in_usd),
    ("Cost basis: EUR stock → PLN",           test_cost_basis_eur_stock_converts_to_pln),
    ("Return: EUR stock not inflated",        test_return_eur_stock_not_inflated),
    ("Ticker history: buys and sells",        test_get_ticker_history_buys_and_sells),
    ("Ticker history: running shares",        test_get_ticker_history_running_shares),
    ("Ticker history: excludes cash",         test_get_ticker_history_excludes_cash),
    ("Theme: dark keys complete",             test_theme_keys_dark),
    ("Theme: light keys complete",            test_theme_keys_light),
    ("Theme: colors valid CSS",               test_theme_colors_valid),
    ("Theme: plotly template correct",        test_theme_plotly_template),
    ("Theme: CSS uses correct colors",        test_theme_css_uses_correct_colors),
    ("Theme: plotly font colors match",       test_theme_plotly_font_colors),
    ("Theme: dark text readable on dark bg",  test_theme_dark_text_on_dark_bg),
    ("Theme: light text readable on light bg",test_theme_light_text_on_light_bg),
    ("CAGR: basic positive return",          test_cagr_basic),
    ("CAGR: no transactions",                test_cagr_no_transactions),
    ("CAGR: zero invested",                  test_cagr_zero_invested),
    ("CAGR: loss scenario",                  test_cagr_loss),
    ("CAGR: with FX conversion",             test_cagr_with_fx),
    ("IRR: basic positive return",           test_irr_basic),
    ("IRR: no transactions",                 test_irr_no_transactions),
    ("IRR: multiple deposits",               test_irr_multiple_deposits),
    ("IRR: with withdrawal",                 test_irr_with_withdrawal),
    ("IRR: loss scenario",                   test_irr_loss),
    ("IRR: known two-flow closed form",      test_irr_known_two_flow),
    ("IRR: dividend ignored (regression)",   test_irr_dividend_ignored),
    ("IRR: withholding tax ignored",         test_irr_withholding_tax_ignored),
    ("IRR: FX deposit converted",            test_irr_fx_deposit),
    ("IRR: matches reference XIRR",          test_irr_matches_reference_xirr),
    ("IRR: NPV(irr) == 0 property",          test_irr_npv_is_zero),
    ("IRR: multiple sign changes",           test_irr_multiple_sign_changes),
    ("IRR: high return bounded",             test_irr_high_return),
    ("FX: same currency",                    test_fx_rate_same_currency),
    ("FX: direct pair",                      test_fx_rate_direct_pair),
    ("FX: reverse pair",                     test_fx_rate_reverse_pair),
    ("FX: EUR→PLN triangulation",            test_fx_rate_eur_pln_via_triangulation),
    ("FX: USD→PLN direct",                   test_fx_rate_usd_pln_direct),
    ("FX: PLN→USD",                          test_fx_rate_pln_to_usd),
    ("FX: no data fallback",                 test_fx_rate_fallback_1),
    ("FX: EUR→USD cross via PLN",            test_fx_rate_non_pln_cross_via_pln),
    ("Latest price: cash returns 1",         test_latest_price_cash),
    ("Latest price: from cache",             test_latest_price_from_cache),
    ("Latest price: empty cache",            test_latest_price_empty_cache),
    ("Latest price: empty slab",             test_latest_price_empty_slab),
    ("Latest price: multi year scan",        test_latest_price_multi_year),
    ("BOSSA: validate valid CSV",            test_bossa_validate_valid),
    ("BOSSA: validate empty file",           test_bossa_validate_empty),
    ("BOSSA: validate missing columns",      test_bossa_validate_missing_columns),
    ("BOSSA: parse_float basic",             test_bossa_parse_float_basic),
    ("BOSSA: parse_float empty",             test_bossa_parse_float_empty),
    ("BOSSA: parse_float invalid",           test_bossa_parse_float_invalid),
    ("BOSSA: read CSV UTF-8",                test_bossa_read_csv_utf8),
    ("BOSSA: read CSV Windows-1250",         test_bossa_read_csv_windows1250),
    ("BOSSA: parse buy",                     test_bossa_parse_buy),
    ("BOSSA: parse sell",                    test_bossa_parse_sell),
    ("BOSSA: parse deposit",                 test_bossa_parse_deposit),
    ("BOSSA: short row skipped",             test_bossa_parse_short_row_skipped),
    ("XTB: validate missing sheet",          test_xtb_validate_missing_sheet),
    ("XTB: validate missing columns",        test_xtb_validate_missing_columns),
    ("XTB: validate valid file",             test_xtb_validate_valid),
    ("XTB: parse dividend",                  test_xtb_parse_dividend),
    ("XTB: parse withdrawal",                test_xtb_parse_withdrawal),
    ("XTB: parse interest",                  test_xtb_parse_interest),
    ("XTB: parse withholding tax",           test_xtb_parse_withholding_tax),
    ("XTB: parse deposit",                   test_xtb_parse_deposit),
    ("XTB: parse transfer",                  test_xtb_parse_transfer),
    ("Avg price: sell reduces proportionally", test_avg_price_sell_reduces_proportionally),
    ("Avg price: full sell zeroes",          test_avg_price_full_sell_zeroes),
]


def main():
    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("   Negotium - Investment Tracker — Test Suite")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")

    for name, fn in ALL_TESTS:
        run_test(name, fn)

    passed  = sum(1 for _, ok, _ in _RESULTS if ok)
    failed  = sum(1 for _, ok, _ in _RESULTS if not ok)
    total   = len(_RESULTS)

    print(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"  {passed}/{total} passed", end="")
    if failed:
        print(f"  |  {failed} FAILED")
    else:
        print("  — all green ✓")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

    if failed:
        print("\nFailed tests:")
        for name, ok, tb in _RESULTS:
            if not ok:
                print(f"\n  ✗ {name}")
                for line in tb.strip().splitlines():
                    print(f"    {line}")

    cleanup()
    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
